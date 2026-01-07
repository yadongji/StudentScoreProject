#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
限时练成绩导入工具
支持导入限时练考试成绩，识别多sheet结构
"""

import openpyxl
import sqlite3
import re
from datetime import datetime
import os

# 数据库路径 - 使用相对路径
DB_PATH = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'StudentData.db')


def connect_db():
    """连接数据库"""
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def parse_exam_name(filename):
    """从文件名解析考试名称和科目"""
    # 文件名格式: 3.高一一级部物理限时练12.16-成绩排名-学生明细.xlsx
    # 提取: 物理限时练12.16
    match = re.search(r'高一.+?部(.+?)限时练([\d.]+)', filename)
    if match:
        subject = match.group(1).strip()
        date_str = match.group(2).strip()
        # 尝试解析日期 12.16 -> 2024-12-16
        try:
            if '.' in date_str:
                parts = date_str.split('.')
                month = int(parts[0])
                day = int(parts[1])
                exam_date = f"2024-{month:02d}-{day:02d}"
            else:
                exam_date = date_str
        except Exception as e:
            print(f"日期解析警告: {e}")
            exam_date = datetime.now().strftime('%Y-%m-%d')

        exam_name = f"{subject}限时练{date_str}"
        return exam_name, subject, exam_date

    return None, None, None


def find_subject_id(conn, subject_name):
    """查找科目ID"""
    cursor = conn.cursor()
    cursor.execute("SELECT SubjectId FROM Subjects WHERE SubjectName = ?", (subject_name,))
    result = cursor.fetchone()
    if result:
        return result['SubjectId']

    # 如果科目不存在，返回None
    return None


def create_time_limit_exam_if_not_exists(conn, exam_name, subject_name, subject_id, exam_date, grade_name='高一'):
    """创建限时练考试（如果不存在）"""
    cursor = conn.cursor()
    cursor.execute("SELECT ExamId FROM TimeLimitExams WHERE ExamName = ?", (exam_name,))
    result = cursor.fetchone()

    if result:
        return result['ExamId']

    # 确定学期
    try:
        month = int(exam_date.split('-')[1])
    except:
        month = 1
    term = '上学期' if month <= 7 else '下学期'

    # 确定学年
    try:
        year = int(exam_date.split('-')[0])
    except:
        year = 2024
    academic_year = f"{year}-{year+1}"

    # 创建新限时练考试
    cursor.execute("""
        INSERT INTO TimeLimitExams (ExamName, ExamDate, SubjectName, SubjectId, GradeName, Term, AcademicYear, Description)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
    """, (exam_name, exam_date, subject_name, subject_id, grade_name, term, academic_year, f"限时练考试"))
    conn.commit()
    return cursor.lastrowid


def clean_student_name(name):
    """清理学生姓名（去除后面的数字）"""
    if not name:
        return None

    # 去除后面的数字（如：张三1 -> 张三）
    name = str(name).strip()
    name = re.sub(r'\d+$', '', name)
    return name if name else None


def detect_sheet_columns(ws):
    """
    智能检测Excel表的列位置
    优先级：
    1. 检测学号列（包含"学号"二字）
    2. 如果没有学号，检测姓名列（包含"姓名"二字）
    3. 检测学科成绩列（学科名下方的"成绩"）
    4. 检测排名列（成绩列下方的"班级排名"/"班次"/"年级排名"）
    """
    # 获取第2行和第3行的标题
    header_row1 = [ws.cell(2, col).value for col in range(1, ws.max_column + 1)]
    header_row2 = [ws.cell(3, col).value for col in range(1, ws.max_column + 1)]

    # 列索引（从1开始）
    student_number_col = None
    student_name_col = None
    score_col = None
    class_rank_col = None
    grade_rank_col = None

    # 1. 优先检测学号列（在第2行或第3行中查找包含"学号"的列）
    for col_idx in range(1, ws.max_column + 1):
        header1 = header_row1[col_idx - 1]
        header2 = header_row2[col_idx - 1]
        if (header1 and '学号' in str(header1)) or (header2 and '学号' in str(header2)):
            student_number_col = col_idx
            break

    # 2. 如果没找到学号，检测姓名列
    if not student_number_col:
        for col_idx in range(1, ws.max_column + 1):
            header1 = header_row1[col_idx - 1]
            header2 = header_row2[col_idx - 1]
            if (header1 and '姓名' in str(header1)) or (header2 and '姓名' in str(header2)):
                student_name_col = col_idx
                break

    # 3. 检测学科成绩列（查找"成绩"字样的列）
    # 查找第3行中包含"成绩"的列
    for col_idx in range(1, ws.max_column + 1):
        header2 = header_row2[col_idx - 1]
        if header2 and '成绩' in str(header2):
            score_col = col_idx
            break

    # 4. 检测排名列
    # 如果找到了成绩列，检查成绩列附近是否有排名列
    if score_col:
        # 检查成绩列后面几列是否有"班级排名"/"班次"
        for col_idx in range(score_col + 1, min(score_col + 3, ws.max_column + 1)):
            header2 = header_row2[col_idx - 1]
            if header2 and ('班级排名' in str(header2) or '班次' in str(header2)):
                class_rank_col = col_idx
                break

        # 检查成绩列后面几列是否有"年级排名"
        for col_idx in range(score_col + 1, min(score_col + 5, ws.max_column + 1)):
            header2 = header_row2[col_idx - 1]
            if header2 and '年级排名' in str(header2):
                grade_rank_col = col_idx
                break

    return student_number_col, student_name_col, score_col, class_rank_col, grade_rank_col


def import_time_limit_sheet(conn, ws, exam_id, subject_id, class_name):
    """导入一个sheet的数据"""
    success_count = 0
    fail_count = 0
    errors = []

    try:
        # 智能检测列位置
        student_number_col, student_name_col, score_col, class_rank_col, grade_rank_col = detect_sheet_columns(ws)

        print(f"\n  检测到的列位置:")
        if student_number_col:
            print(f"    学号列: 第{student_number_col}列")
        if student_name_col:
            print(f"    姓名列: 第{student_name_col}列")
        if score_col:
            print(f"    成绩列: 第{score_col}列")
        if class_rank_col:
            print(f"    班级排名列: 第{class_rank_col}列")
        if grade_rank_col:
            print(f"    年级排名列: 第{grade_rank_col}列")

        # 验证是否检测到必要的列
        if not score_col:
            errors.append("未检测到成绩列，请检查Excel格式")
            fail_count += 1
            return success_count, fail_count, errors

        if not student_number_col and not student_name_col:
            errors.append("未检测到学号列或姓名列，请检查Excel格式")
            fail_count += 1
            return success_count, fail_count, errors

        # 从第4行开始读取数据
        for row_idx in range(4, ws.max_row + 1):
            try:
                row = list(ws[row_idx])

                # 读取学生标识（学号优先，其次姓名）
                school_number = None
                name = None

                if student_number_col:
                    number_cell = row[student_number_col - 1]
                    school_number = str(number_cell.value).strip() if number_cell.value else None

                if student_name_col:
                    name_cell = row[student_name_col - 1]
                    name = clean_student_name(name_cell.value)

                # 如果有学号，使用学号；否则使用姓名
                if not school_number and not name:
                    fail_count += 1
                    continue

                # 读取成绩
                score_cell = row[score_col - 1]
                score_value = str(score_cell.value).strip() if score_cell.value else ''
                is_absent = False

                if score_value in ['--', '-', '']:
                    is_absent = True
                    score = None
                else:
                    try:
                        score = float(score_value)
                    except:
                        is_absent = True
                        score = None

                # 查找学生ID
                cursor = conn.cursor()
                student_result = None

                # 优先使用学号查询
                if school_number:
                    cursor.execute("SELECT StudentId FROM Students WHERE StudentNumber = ?", (school_number,))
                    student_result = cursor.fetchone()

                # 如果学号查询失败，尝试用姓名和班级匹配
                if not student_result and name:
                    cursor.execute("SELECT StudentId FROM Students WHERE StudentName = ? AND ClassName = ?",
                                 (name, class_name))
                    student_result = cursor.fetchone()

                if not student_result:
                    errors.append(f"找不到学生: {name or '未知'}({school_number or '无学号'})")
                    fail_count += 1
                    continue

                student_id = student_result['StudentId']

                # 读取班级排名
                class_rank = None
                if class_rank_col:
                    class_rank_cell = row[class_rank_col - 1]
                    if class_rank_cell.value is not None:
                        rank_value = str(class_rank_cell.value).strip()
                        if rank_value and rank_value not in ['--', '-', '', 'None']:
                            try:
                                class_rank = int(float(rank_value))
                            except:
                                try:
                                    class_rank = int(rank_value)
                                except:
                                    pass

                # 读取年级排名
                grade_rank = None
                if grade_rank_col:
                    grade_rank_cell = row[grade_rank_col - 1]
                    if grade_rank_cell.value is not None:
                        rank_value = str(grade_rank_cell.value).strip()
                        if rank_value and rank_value not in ['--', '-', '', 'None']:
                            try:
                                grade_rank = int(float(rank_value))
                            except:
                                try:
                                    grade_rank = int(rank_value)
                                except:
                                    pass

                # 如果缺考，记录但不保存到数据库
                if is_absent:
                    errors.append(f"{name or '未知'}({school_number or '无学号'}): 缺考")
                    fail_count += 1
                    continue

                # 检查是否已存在（在TimeLimitScores表中使用TimeLimitExamId）
                cursor.execute("""
                    SELECT ScoreId FROM TimeLimitScores
                    WHERE TimeLimitExamId = ? AND StudentId = ? AND SubjectId = ?
                """, (exam_id, student_id, subject_id))

                if cursor.fetchone():
                    # 更新
                    cursor.execute("""
                        UPDATE TimeLimitScores
                        SET Score = ?, ClassRank = ?, GradeRank = ?, UpdatedAt = datetime('now')
                        WHERE TimeLimitExamId = ? AND StudentId = ? AND SubjectId = ?
                    """, (score, class_rank, grade_rank, exam_id, student_id, subject_id))
                else:
                    # 插入
                    cursor.execute("""
                        INSERT INTO TimeLimitScores (TimeLimitExamId, StudentId, SubjectId, Score, ClassRank, GradeRank)
                        VALUES (?, ?, ?, ?, ?, ?)
                    """, (exam_id, student_id, subject_id, score, class_rank, grade_rank))

                conn.commit()
                success_count += 1

            except Exception as e:
                fail_count += 1
                errors.append(f"第{row_idx}行错误: {str(e)}")

    except Exception as e:
        errors.append(f"Sheet读取错误: {str(e)}")

    return success_count, fail_count, errors


def import_time_limit_excel(excel_path, grade_name='高一'):
    """导入限时练Excel文件"""
    print("=" * 80)
    print("限时练成绩导入工具")
    print("=" * 80)

    # 解析考试信息
    filename = excel_path.split('\\')[-1]
    exam_name, subject_name, exam_date = parse_exam_name(filename)

    if not exam_name or not subject_name:
        print(f"❌ 无法从文件名解析考试信息: {filename}")
        return

    print(f"\n考试名称: {exam_name}")
    print(f"科目: {subject_name}")
    print(f"日期: {exam_date}")
    print(f"文件: {filename}")
    print(f"年级: {grade_name}")

    conn = connect_db()

    try:
        # 查找科目ID
        subject_id = find_subject_id(conn, subject_name)
        if not subject_id:
            print(f"❌ 科目不存在: {subject_name}")
            print("   请先在数据库中创建该科目")
            return

        # 创建限时练考试
        exam_id = create_time_limit_exam_if_not_exists(conn, exam_name, subject_name, subject_id, exam_date, grade_name)
        print(f"考试ID: {exam_id}")

        # 加载Excel
        print(f"\n正在加载Excel文件...")
        wb = openpyxl.load_workbook(excel_path, data_only=True)

        total_success = 0
        total_fail = 0
        all_errors = []

        # 遍历所有sheet
        for sheet_name in wb.sheetnames:
            print(f"\n{'='*80}")
            print(f"处理Sheet: {sheet_name}")
            print(f"{'='*80}")

            ws = wb[sheet_name]

            # Sheet名称即为班级名称
            class_name = sheet_name.strip()
            print(f"班级: {class_name}")

            success, fail, errors = import_time_limit_sheet(conn, ws, exam_id, subject_id, class_name)

            total_success += success
            total_fail += fail
            all_errors.extend(errors)

            print(f"\n导入完成: 成功 {success} 条, 失败 {fail} 条")

            # 统计缺考人数
            absent_count = sum(1 for err in errors if '缺考' in err)

            if errors:
                print(f"处理详情:")
                if absent_count > 0:
                    print(f"  缺考: {absent_count} 人")

                # 显示其他错误
                other_errors = [err for err in errors if '缺考' not in err]
                if other_errors:
                    print(f"  错误:")
                    for err in other_errors[:10]:  # 只显示前10个错误
                        print(f"    - {err}")
                    if len(other_errors) > 10:
                        print(f"    ... 还有 {len(other_errors) - 10} 个错误")

        wb.close()

        print(f"\n{'='*80}")
        print(f"总计导入完成: 成功 {total_success} 条, 失败 {total_fail} 条")

        # 统计总缺考人数
        total_absent = sum(1 for err in all_errors if '缺考' in err)
        if total_absent > 0:
            print(f"缺考人数: {total_absent} 人")

        print(f"{'='*80}")

    except Exception as e:
        print(f"\n❌ 发生错误: {e}")
        import traceback
        traceback.print_exc()

    finally:
        conn.close()


def main():
    import sys
    import os

    # 第一次导入
    if len(sys.argv) > 1:
        excel_path = sys.argv[1]
        grade_name = sys.argv[2] if len(sys.argv) > 2 else '高一'
        import_time_limit_excel(excel_path, grade_name)
    else:
        excel_path = input("请输入Excel文件路径: ").strip()
        grade_name = input("请输入年级（默认：高一）: ").strip() or '高一'
        import_time_limit_excel(excel_path, grade_name)

    # 循环导入，直到用户选择退出
    while True:
        print("\n" + "=" * 80)
        print("是否继续导入其他文件？")
        print("  1. 继续导入（拖拽Excel文件到此窗口或输入路径）")
        print("  2. 退出")
        print("=" * 80)

        choice = input("\n请选择（1-2，默认2）: ").strip()
        if not choice or choice == '2':
            print("\n✅ 导入程序结束")
            break

        if choice == '1':
            # 尝试从剪贴板或输入获取文件路径
            print("\n请拖拽Excel文件到此窗口，或输入文件路径:")
            new_excel_path = input().strip()

            # 去除可能包含的引号
            new_excel_path = new_excel_path.strip('"').strip("'")

            if not new_excel_path:
                print("❌ 未提供文件路径，跳过")
                continue

            if not os.path.exists(new_excel_path):
                print(f"❌ 文件不存在: {new_excel_path}")
                continue

            if not new_excel_path.lower().endswith('.xlsx'):
                print("❌ 文件格式错误，请选择.xlsx文件")
                continue

            # 询问是否修改年级
            change_grade = input(f"是否修改年级（当前：{grade_name}）？(y/N): ").strip().lower()
            if change_grade == 'y':
                new_grade = input("请输入年级: ").strip()
                if new_grade:
                    grade_name = new_grade

            # 导入新文件
            import_time_limit_excel(new_excel_path, grade_name)
        else:
            print("❌ 无效选择")



if __name__ == '__main__':
    main()
