#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
班级排名趋势分析工具
查看整个班级每个学生的排名变化情况
"""

import sqlite3
from datetime import datetime, timedelta
import os
import sys

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# 数据库路径 - 使用相对路径
DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'StudentData.db')

# 科目映射
SUBJECT_MAPPING = {
    1: '语文', 2: '数学', 3: '英语',
    4: '物理', 5: '化学', 6: '生物',
    7: '政治', 8: '历史', 9: '地理',
    10: '总分'
}


def connect_db():
    """连接数据库"""
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def get_all_classes(conn):
    """获取所有班级"""
    cursor = conn.cursor()
    cursor.execute("""
        SELECT DISTINCT ClassName
        FROM Students
        WHERE ClassName IS NOT NULL AND ClassName != ''
        ORDER BY ClassName
    """)
    return cursor.fetchall()


def get_students_in_class(conn, class_name):
    """获取班级所有学生"""
    cursor = conn.cursor()
    cursor.execute("""
        SELECT StudentId, StudentNumber, StudentName
        FROM Students
        WHERE ClassName = ?
        ORDER BY StudentNumber
    """, (class_name,))
    return cursor.fetchall()


def get_class_rank_trend(conn, class_name, subject_id, start_date=None, end_date=None, rank_type='grade'):
    """获取班级某科目的排名趋势

    Args:
        conn: 数据库连接
        class_name: 班级名称
        subject_id: 科目ID
        start_date: 开始日期
        end_date: 结束日期
        rank_type: 排名类型，'class'为班级排名，'grade'为年级排名（默认）
    """
    rank_field = 's.ClassRank' if rank_type == 'class' else 's.GradeRank'

    sql = f"""
        SELECT
            e.ExamId,
            e.ExamName,
            e.ExamDate,
            s.StudentId,
            st.StudentNumber,
            st.StudentName,
            {rank_field} as Rank
        FROM Scores s
        JOIN Exams e ON s.ExamId = e.ExamId
        JOIN Students st ON s.StudentId = st.StudentId
        WHERE st.ClassName = ? AND s.SubjectId = ?
    """

    params = [class_name, subject_id]

    if start_date:
        sql += " AND e.ExamDate >= ?"
        params.append(start_date)

    if end_date:
        sql += " AND e.ExamDate <= ?"
        params.append(end_date)

    sql += " ORDER BY e.ExamDate, st.StudentNumber"

    cursor = conn.cursor()
    cursor.execute(sql, params)
    return cursor.fetchall()


def get_all_subjects_with_scores(conn, class_name):
    """获取班级有成绩记录的科目"""
    cursor = conn.cursor()
    cursor.execute("""
        SELECT DISTINCT
            sb.SubjectId,
            sb.SubjectName,
            sb.SortOrder,
            COUNT(s.ScoreId) as RecordCount
        FROM Subjects sb
        JOIN Scores s ON sb.SubjectId = s.SubjectId
        JOIN Students st ON s.StudentId = st.StudentId
        WHERE st.ClassName = ?
        GROUP BY sb.SubjectId, sb.SubjectName, sb.SortOrder
        ORDER BY sb.SortOrder
    """, (class_name,))
    return cursor.fetchall()


def format_name(name):
    """格式化姓名：两个字中间加空格"""
    if len(name) == 2:
        return f"{name[0]}  {name[1]}"
    return name


def print_class_rank_summary(scores, class_name, subject_name, decline_threshold=5, rank_type='grade'):
    """打印班级排名变化摘要

    Args:
        scores: 成绩数据
        class_name: 班级名称
        subject_name: 科目名称
        decline_threshold: 退步显示阈值
        rank_type: 排名类型，'class'为班级排名，'grade'为年级排名
    """
    if not scores:
        print(f"⚠️  {class_name}没有{subject_name}成绩记录")
        return

    rank_type_name = "年级排名" if rank_type == 'grade' else "班级排名"

    # 按学生分组
    students_data = {}
    for s in scores:
        if s['StudentId'] not in students_data:
            students_data[s['StudentId']] = {
                'name': s['StudentName'],
                'number': s['StudentNumber'],
                'records': []
            }
        students_data[s['StudentId']]['records'].append(s)

    # 调试信息：显示有多少学生有记录
    print(f"\n📋 调试信息: 共获取 {len(scores)} 条成绩记录，涉及 {len(students_data)} 名学生")

    # 找出进步和退步最大的学生
    improvements = []
    declines = []
    no_change = []

    # 统计每个学生的考试次数
    exam_count_distribution = {}
    for student_id, data in students_data.items():
        records = sorted(data['records'], key=lambda x: datetime.strptime(x['ExamDate'], '%Y-%m-%d'))
        exam_count = len(records)
        exam_count_distribution[exam_count] = exam_count_distribution.get(exam_count, 0) + 1

    # 显示考试次数分布
    print(f"📊 考试次数分布:", end=" ")
    for exam_count in sorted(exam_count_distribution.keys()):
        print(f"{exam_count}次考试: {exam_count_distribution[exam_count]}人", end="; ")
    print()

    for student_id, data in students_data.items():
        records = sorted(data['records'], key=lambda x: datetime.strptime(x['ExamDate'], '%Y-%m-%d'))

        if len(records) >= 2:
            first_rank = records[0]['Rank']
            last_rank = records[-1]['Rank']

            # 只有当第一次和最后一次排名都存在时，才计算进步/退步
            if first_rank is not None and last_rank is not None:
                change = first_rank - last_rank  # 正数表示进步（排名上升）

                if change > 0:
                    improvements.append((change, data['name'], data['number'],
                                       records[0]['ExamName'], records[0]['Rank'],
                                       records[-1]['ExamName'], records[-1]['Rank']))
                elif change < 0:
                    declines.append((abs(change), data['name'], data['number'],
                                   records[0]['ExamName'], records[0]['Rank'],
                                   records[-1]['ExamName'], records[-1]['Rank']))
                else:
                    no_change.append((data['name'], data['number']))

    improvements.sort(reverse=True)
    declines.sort(reverse=True)

    print(f"\n{'='*102}")
    print(f"📊 {class_name} - {subject_name}{rank_type_name}变化分析")
    print(f"{'='*102}")

    # 显示所有进步的学生
    if improvements:
        print(f"\n📈 进步学生（共{len(improvements)}人）:")
        header = f"{'序号':^6} {'姓名':>9} {'学号':^13} {'变化':^8} {'初始考试':^20} {'初始排名':^8} {'最近考试':^20} {'最新排名':^8}"
        print(header)
        print(f"{'-'*102}")
        for i, (change, name, number, first_exam, first_rank, last_exam, last_rank) in enumerate(improvements, 1):
            formatted_name = format_name(name)
            first_rank_str = str(first_rank) if first_rank else 'N/A'
            last_rank_str = str(last_rank) if last_rank else 'N/A'
            number_str = number or ''
            print(f"{i:^6} {formatted_name:>9} {number_str:^13} +{change:^6} {first_exam:^20} {first_rank_str:^8} {last_exam:^20} {last_rank_str:^8}")

    # 只显示退步明显的学生（超过阈值）
    if declines:
        significant_declines = [d for d in declines if d[0] >= decline_threshold]
        if significant_declines:
            print(f"\n📉 明显退步学生（退步{decline_threshold}名及以上，共{len(significant_declines)}人）:")
            header = f"{'序号':^6} {'姓名':>9} {'学号':^13} {'变化':^8} {'初始考试':^20} {'初始排名':^8} {'最近考试':^20} {'最新排名':^8}"
            print(header)
            print(f"{'-'*102}")
            for i, (change, name, number, first_exam, first_rank, last_exam, last_rank) in enumerate(significant_declines, 1):
                formatted_name = format_name(name)
                first_rank_str = str(first_rank) if first_rank else 'N/A'
                last_rank_str = str(last_rank) if last_rank else 'N/A'
                number_str = number or ''
                print(f"{i:^6} {formatted_name:>9} {number_str:^13} -{change:^6} {first_exam:^20} {first_rank_str:^8} {last_exam:^20} {last_rank_str:^8}")
        else:
            print(f"\n📉 没有学生退步{decline_threshold}名及以上")

    # 显示统计信息
    total_students = len(students_data)
    valid_students = len(improvements) + len(declines) + len(no_change)
    incomplete_students = total_students - valid_students  # 排名缺失的学生
    print(f"\n📌 统计摘要:")
    print(f"  班级总人数: {total_students}人")
    print(f"  有效记录: {valid_students}人")
    if incomplete_students > 0:
        print(f"  排名缺失: {incomplete_students}人（某次考试排名为空）")
    print(f"  进步人数: {len(improvements)}人")
    print(f"  退步人数: {len(declines)}人")
    print(f"  保持不变: {len(no_change)}人")

    if improvements:
        avg_improvement = sum(i[0] for i in improvements) / len(improvements)
        print(f"  平均进步: {avg_improvement:.1f}名")

    if declines:
        avg_decline = sum(d[0] for d in declines) / len(declines)
        print(f"  平均退步: {avg_decline:.1f}名")

    print(f"{'='*102}")


def print_class_rank_sum_trend(scores, class_name, subject_name):
    """打印班级学科排名总和变化

    Args:
        scores: 成绩数据
        class_name: 班级名称
        subject_name: 科目名称
    """
    if not scores:
        print(f"⚠️  {class_name}没有{subject_name}成绩记录")
        return

    # 按考试分组，计算每次考试的所有学生排名总和
    exam_rank_sums = {}
    exam_students_count = {}

    for s in scores:
        exam_name = s['ExamName']
        exam_date = s['ExamDate']
        rank = s['Rank']

        if exam_name not in exam_rank_sums:
            exam_rank_sums[exam_name] = {
                'ExamDate': exam_date,
                'RankSum': 0,
                'StudentCount': 0
            }

        if rank is not None:
            exam_rank_sums[exam_name]['RankSum'] += rank
            exam_rank_sums[exam_name]['StudentCount'] += 1

    if not exam_rank_sums:
        print(f"⚠️  没有有效的排名数据")
        return

    # 按日期排序
    sorted_exams = sorted(exam_rank_sums.items(), key=lambda x: datetime.strptime(x[1]['ExamDate'], '%Y-%m-%d'))

    if len(sorted_exams) < 2:
        print(f"⚠️  该班级{subject_name}只有{len(sorted_exams)}次考试记录，无法计算排名总和变化")
        return

    # 计算排名总和变化
    rank_sum_changes = []
    for i in range(1, len(sorted_exams)):
        prev_exam_name, prev_exam_data = sorted_exams[i-1]
        curr_exam_name, curr_exam_data = sorted_exams[i]

        prev_rank_sum = prev_exam_data['RankSum']
        curr_rank_sum = curr_exam_data['RankSum']
        prev_student_count = prev_exam_data['StudentCount']
        curr_student_count = curr_exam_data['StudentCount']

        change = prev_rank_sum - curr_rank_sum  # 正数表示进步（排名总和变小）

        if change > 0:
            trend = "📈 进步"
            trend_icon = "↑"
        elif change < 0:
            trend = "📉 倒退"
            trend_icon = "↓"
        else:
            trend = "➡️ 持平"
            trend_icon = "→"

        rank_sum_changes.append({
            'prev_exam': prev_exam_name,
            'curr_exam': curr_exam_name,
            'prev_date': prev_exam_data['ExamDate'],
            'curr_date': curr_exam_data['ExamDate'],
            'prev_rank_sum': prev_rank_sum,
            'curr_rank_sum': curr_rank_sum,
            'change': change,
            'trend': trend,
            'trend_icon': trend_icon,
            'prev_count': prev_student_count,
            'curr_count': curr_student_count
        })

    print(f"\n{'='*100}")
    print(f"📊 {class_name} - {subject_name}年级排名总和变化分析")
    print(f"{'='*100}")

    # 显示变化表格
    print(f"\n{'序号':^6} {'上次考试':^18} {'上次排名总和':^12} {'当前考试':^18} {'当前排名总和':^12} {'变化':^10} {'趋势'}")
    print(f"{'-'*100}")

    for i, change_data in enumerate(rank_sum_changes, 1):
        prev_exam = change_data['prev_exam']
        curr_exam = change_data['curr_exam']
        prev_rank_sum = change_data['prev_rank_sum']
        curr_rank_sum = change_data['curr_rank_sum']
        change = change_data['change']
        trend = change_data['trend']
        trend_icon = change_data['trend_icon']
        prev_count = change_data['prev_count']
        curr_count = change_data['curr_count']

        change_str = f"+{change}" if change > 0 else f"{change}"

        print(f"{i:^6} {prev_exam:^18} {prev_rank_sum:^12} {curr_exam:^18} {curr_rank_sum:^12} {change_str:^10} {trend}")

    # 统计摘要
    total_improvements = sum(1 for c in rank_sum_changes if c['change'] > 0)
    total_declines = sum(1 for c in rank_sum_changes if c['change'] < 0)
    total_no_change = sum(1 for c in rank_sum_changes if c['change'] == 0)

    total_change = sum(c['change'] for c in rank_sum_changes)
    avg_change = total_change / len(rank_sum_changes) if rank_sum_changes else 0

    print(f"\n📌 统计摘要:")
    print(f"  考试次数: {len(sorted_exams)}次")
    print(f"  进步次数: {total_improvements}次")
    print(f"  倒退次数: {total_declines}次")
    print(f"  持平次数: {total_no_change}次")
    print(f"  总体变化: {total_change:+.0f}")
    print(f"  平均变化: {avg_change:+.1f}")

    if total_change > 0:
        print(f"  ✅ 总体趋势：进步")
    elif total_change < 0:
        print(f"  ⚠️  总体趋势：倒退")
    else:
        print(f"  ➡️ 总体趋势：持平")

    print(f"{'='*100}")


def get_student_rank_changes(scores):
    """分析学生进步/退退情况

    Args:
        scores: 成绩数据

    Returns:
        tuple: (improvements, declines, no_change, first_exam_name, last_exam_name)
    """
    if not scores:
        return [], [], [], None, None

    # 按学生分组
    students_data = {}
    for s in scores:
        if s['StudentId'] not in students_data:
            students_data[s['StudentId']] = {
                'name': s['StudentName'],
                'number': s['StudentNumber'],
                'records': []
            }
        students_data[s['StudentId']]['records'].append(s)

    # 找出进步和退退的学生
    improvements = []
    declines = []
    no_change = []

    for student_id, data in students_data.items():
        records = sorted(data['records'], key=lambda x: datetime.strptime(x['ExamDate'], '%Y-%m-%d'))

        if len(records) >= 2:
            first_rank = records[0]['Rank']
            last_rank = records[-1]['Rank']

            # 只有当第一次和最后一次排名都存在时，才计算进步/退步
            if first_rank is not None and last_rank is not None:
                change = first_rank - last_rank  # 正数表示进步（排名上升）

                if change > 0:
                    improvements.append({
                        'name': data['name'],
                        'number': data['number'],
                        'change': change,
                        'first_exam': records[0]['ExamName'],
                        'first_rank': records[0]['Rank'],
                        'last_exam': records[-1]['ExamName'],
                        'last_rank': records[-1]['Rank']
                    })
                elif change < 0:
                    declines.append({
                        'name': data['name'],
                        'number': data['number'],
                        'change': abs(change),
                        'first_exam': records[0]['ExamName'],
                        'first_rank': records[0]['Rank'],
                        'last_exam': records[-1]['ExamName'],
                        'last_rank': records[-1]['Rank']
                    })
                else:
                    no_change.append({
                        'name': data['name'],
                        'number': data['number']
                    })

    # 排序
    improvements.sort(key=lambda x: x['change'], reverse=True)
    declines.sort(key=lambda x: x['change'], reverse=True)

    # 获取第一次和最后一次考试名称
    if scores:
        sorted_scores = sorted(scores, key=lambda x: datetime.strptime(x['ExamDate'], '%Y-%m-%d'))
        first_exam_name = sorted_scores[0]['ExamName']
        last_exam_name = sorted_scores[-1]['ExamName']
    else:
        first_exam_name = None
        last_exam_name = None

    return improvements, declines, no_change, first_exam_name, last_exam_name


def generate_excel_report(improvements, declines, class_name, first_exam, last_exam, subject_name):
    """生成Excel报告

    Args:
        improvements: 进步学生列表
        declines: 退步学生列表
        class_name: 班级名称
        first_exam: 开始考试名称
        last_exam: 结束考试名称
        subject_name: 科目名称
    """
    if not HAS_OPENPYXL:
        print("❌ 缺少openpyxl库，无法生成Excel文件")
        print("请运行: pip install openpyxl")
        return False

    try:
        # 创建工作簿
        wb = Workbook()

        # 删除默认的sheet
        wb.remove(wb.active)

        # 定义样式
        header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        cell_alignment = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 创建进步sheet
        if improvements:
            ws_improve = wb.create_sheet(title='进步学生')
            # 写入表头
            headers = ['序号', '姓名', '学号', '进步名次', '初始考试', '初始排名', '最近考试', '最新排名']
            for col, header in enumerate(headers, 1):
                cell = ws_improve.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

            # 写入数据
            for idx, student in enumerate(improvements, 2):
                data = [
                    idx - 1,
                    student['name'],
                    student['number'],
                    student['change'],
                    student['first_exam'],
                    student['first_rank'],
                    student['last_exam'],
                    student['last_rank']
                ]
                for col, value in enumerate(data, 1):
                    cell = ws_improve.cell(row=idx, column=col, value=value)
                    cell.alignment = cell_alignment
                    cell.border = thin_border

            # 调整列宽
            ws_improve.column_dimensions['A'].width = 6
            ws_improve.column_dimensions['B'].width = 10
            ws_improve.column_dimensions['C'].width = 15
            ws_improve.column_dimensions['D'].width = 10
            ws_improve.column_dimensions['E'].width = 20
            ws_improve.column_dimensions['F'].width = 10
            ws_improve.column_dimensions['G'].width = 20
            ws_improve.column_dimensions['H'].width = 10

            # 冻结首行
            ws_improve.freeze_panes = 'A2'
        else:
            ws_improve = wb.create_sheet(title='进步学生')
            ws_improve.cell(row=1, column=1, value='无进步学生记录')

        # 创建退退sheet
        if declines:
            ws_decline = wb.create_sheet(title='退步学生')
            # 写入表头
            headers = ['序号', '姓名', '学号', '退步名次', '初始考试', '初始排名', '最近考试', '最新排名']
            for col, header in enumerate(headers, 1):
                cell = ws_decline.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
                cell.alignment = header_alignment
                cell.border = thin_border

            # 写入数据
            for idx, student in enumerate(declines, 2):
                data = [
                    idx - 1,
                    student['name'],
                    student['number'],
                    student['change'],
                    student['first_exam'],
                    student['first_rank'],
                    student['last_exam'],
                    student['last_rank']
                ]
                for col, value in enumerate(data, 1):
                    cell = ws_decline.cell(row=idx, column=col, value=value)
                    cell.alignment = cell_alignment
                    cell.border = thin_border

            # 调整列宽
            ws_decline.column_dimensions['A'].width = 6
            ws_decline.column_dimensions['B'].width = 10
            ws_decline.column_dimensions['C'].width = 15
            ws_decline.column_dimensions['D'].width = 10
            ws_decline.column_dimensions['E'].width = 20
            ws_decline.column_dimensions['F'].width = 10
            ws_decline.column_dimensions['G'].width = 20
            ws_decline.column_dimensions['H'].width = 10

            # 冻结首行
            ws_decline.freeze_panes = 'A2'
        else:
            ws_decline = wb.create_sheet(title='退步学生')
            ws_decline.cell(row=1, column=1, value='无退步学生记录')

        # 生成文件名：班级+开始的考试+当前的考试名字
        filename = f"{class_name}_{first_exam}_to_{last_exam}_{subject_name}.xlsx"

        # 保存到桌面
        desktop = os.path.join(os.path.expanduser('~'), 'Desktop')
        file_path = os.path.join(desktop, filename)

        wb.save(file_path)
        print(f"\n✅ Excel报告已生成: {file_path}")
        print(f"  - 进步学生: {len(improvements)}人")
        print(f"  - 退步学生: {len(declines)}人")

        return True

    except Exception as e:
        print(f"\n❌ 生成Excel失败: {e}")
        import traceback
        traceback.print_exc()
        return False


def analyze_and_export_excel(conn, class_name, subject_id, subject_name):
    """分析班级排名变化并生成Excel报告

    Args:
        conn: 数据库连接
        class_name: 班级名称
        subject_id: 科目ID
        subject_name: 科目名称
    """
    print(f"\n{'='*102}")
    print(f"📊 {class_name} - {subject_name} 排名变化分析")
    print(f"{'='*102}")

    # 获取年级排名数据
    scores = get_class_rank_trend(conn, class_name, subject_id, None, None, 'grade')

    if not scores:
        print(f"⚠️  {class_name}没有{subject_name}成绩记录")
        return

    # 分析进步/退退情况
    improvements, declines, no_change, first_exam, last_exam = get_student_rank_changes(scores)

    print(f"\n📋 分析时间范围: {first_exam} 至 {last_exam}")
    print(f"📈 进步学生: {len(improvements)}人")
    print(f"📉 退步学生: {len(declines)}人")
    print(f"➡️  持平学生: {len(no_change)}人")

    # 显示前10名进步学生
    if improvements:
        print(f"\n📈 进步学生Top{min(10, len(improvements))}:")
        for i, student in enumerate(improvements[:10], 1):
            print(f"  {i}. {student['name']} (学号: {student['number']}) - 进步{student['change']}名 "
                  f"({student['first_rank']}名 → {student['last_rank']}名)")

    # 显示前10名退步学生
    if declines:
        print(f"\n📉 退步学生Top{min(10, len(declines))}:")
        for i, student in enumerate(declines[:10], 1):
            print(f"  {i}. {student['name']} (学号: {student['number']}) - 退步{student['change']}名 "
                  f"({student['first_rank']}名 → {student['last_rank']}名)")

    # 询问是否生成Excel
    if improvements or declines:
        while True:
            export_choice = input(f"\n是否生成Excel报告？(y/n): ").strip().lower()
            if export_choice in ['y', 'n']:
                break
            print("❌ 请输入 y 或 n")

        if export_choice == 'y':
            generate_excel_report(improvements, declines, class_name, first_exam, last_exam, subject_name)
    else:
        print("\n⚠️  没有进步或退步的学生记录，无需生成Excel报告")


def main():
    print("=" * 100)
    print("                              班级排名趋势分析工具")
    print("=" * 100)

    conn = connect_db()

    try:
        while True:  # 主循环：支持连续查询
            # 第一步：选择班级
            classes = get_all_classes(conn)

            if not classes:
                print("❌ 数据库中没有班级信息")
                return

            print(f"\n找到 {len(classes)} 个班级:")
            for i, c in enumerate(classes, 1):
                print(f"  {i}. {c['ClassName']}")

            choice = input("\n请选择班级编号: ").strip()
            if not choice.isdigit() or int(choice) < 1 or int(choice) > len(classes):
                print("❌ 无效选择")
                continue

            class_name = classes[int(choice) - 1]['ClassName']

            # 第二步：显示可用科目
            subjects = get_all_subjects_with_scores(conn, class_name)

            if not subjects:
                print(f"⚠️  {class_name}没有成绩记录")
                continue

            print(f"\n{'='*100}")
            print(f"  {class_name} 的考试科目")
            print(f"{'='*100}")

            for i, (subject_id, subject_name, sort_order, count) in enumerate(subjects, 1):
                print(f"  {i}. {subject_name} ({count}条记录)")

            choice = input(f"\n请选择科目（1-{len(subjects)}）: ").strip()
            if not choice.isdigit() or int(choice) < 1 or int(choice) > len(subjects):
                print("❌ 无效选择")
                continue

            subject_id, subject_name = subjects[int(choice) - 1][:2]

            # 第三步：选择分析类型
            print(f"\n{'='*100}")
            print(f"  选择分析类型")
            print(f"{'='*100}")
            print(f"  1. 学生个人排名变化分析（现有功能）")
            print(f"  2. 班级学科排名总和变化分析（新功能）")
            print(f"{'='*100}")

            analysis_type_choice = input(f"\n请选择分析类型（1-2，默认1）: ").strip()
            if not analysis_type_choice or not analysis_type_choice.isdigit():
                analysis_type = 1
            else:
                analysis_type = int(analysis_type_choice)

            if analysis_type == 2:
                # 新功能：班级学科排名总和变化分析
                # 直接调用，不需要选择排名类型和时间范围
                scores = get_class_rank_trend(conn, class_name, subject_id, None, None, 'grade')
                print_class_rank_sum_trend(scores, class_name, subject_name)
                continue

            # 原有功能：学生个人排名变化分析
            # 第四步：选择排名类型
            print(f"\n{'='*100}")
            print(f"  选择排名类型")
            print(f"{'='*100}")
            print(f"  1. 年级排名（默认）")
            print(f"  2. 班级排名")

            rank_type_choice = input(f"\n请选择排名类型（1-2，默认1）: ").strip()
            if not rank_type_choice or not rank_type_choice.isdigit():
                rank_type = 'grade'
            elif rank_type_choice == '1':
                rank_type = 'grade'
            else:
                rank_type = 'class'

            rank_type_name = "年级排名" if rank_type == 'grade' else "班级排名"
            print(f"  排名类型: {rank_type_name}")

            # 第五步：选择时间范围
            print(f"\n{'='*100}")
            print(f"  选择时间范围")
            print(f"{'='*100}")
            print(f"  1. 本学期（最近6个月）")
            print(f"  2. 最近两次考试")
            print(f"  3. 全部历史数据")

            choice = input(f"\n请选择时间范围（1-3，默认3）: ").strip()
            if not choice or not choice.isdigit():
                choice = 3
            else:
                choice = int(choice)

            start_date = None
            end_date = None

            if choice == 1:
                # 本学期（最近6个月）
                end_date = datetime.now().strftime('%Y-%m-%d')
                start_date = (datetime.now() - timedelta(days=180)).strftime('%Y-%m-%d')
                print(f"  时间范围: {start_date} 至 {end_date}")
            elif choice == 2:
                # 获取最近两次考试的日期（从 Exams 表中查询该科目有考试的记录）
                cursor = conn.cursor()
                cursor.execute("""
                    SELECT DISTINCT e.ExamId, e.ExamDate, e.ExamName
                    FROM Exams e
                    JOIN Scores s ON e.ExamId = s.ExamId
                    WHERE s.SubjectId = ?
                    ORDER BY e.ExamDate DESC
                    LIMIT 2
                """, (subject_id,))
                recent_exams = cursor.fetchall()

                if len(recent_exams) >= 2:
                    # 获取这两次考试的日期范围
                    exam_dates = [e['ExamDate'] for e in recent_exams]
                    start_date = min(exam_dates)
                    end_date = max(exam_dates)
                    print(f"  时间范围: 最近两次考试 ({start_date} 至 {end_date})")
                    print(f"    考试1: {recent_exams[1]['ExamName']} ({recent_exams[1]['ExamDate']})")
                    print(f"    考试2: {recent_exams[0]['ExamName']} ({recent_exams[0]['ExamDate']})")
                else:
                    print("  ⚠️ 考试次数不足2次，使用全部数据")
            else:
                print("  时间范围: 全部历史数据")

            # 第六步：选择退步阈值
            decline_threshold = input("\n退步显示阈值（退步多少名以上才显示，默认5名）: ").strip()
            if decline_threshold and decline_threshold.isdigit():
                decline_threshold = int(decline_threshold)
            else:
                decline_threshold = 5
            print(f"退步阈值: {decline_threshold}名")

            # 第七步：获取数据并分析
            print(f"\n正在获取{class_name}的{subject_name}{rank_type_name}数据...")
            scores = get_class_rank_trend(conn, class_name, subject_id, start_date, end_date, rank_type)

            if not scores:
                print(f"⚠️  该班级没有{subject_name}成绩记录")
                continue

            # 打印摘要
            print_class_rank_summary(scores, class_name, subject_name, decline_threshold, rank_type)

            # 询问是否继续
            while True:
                continue_query = input("\n是否继续查询？(y/n): ").strip().lower()
                if continue_query in ['y', 'n']:
                    break
                print("❌ 请输入 y 或 n")

            if continue_query == 'n':
                print("\n感谢使用班级排名趋势分析工具，再见！")
                break  # 退出主循环

    except Exception as e:
        print(f"\n❌ 发生错误: {e}")
        import traceback
        traceback.print_exc()

    finally:
        conn.close()


if __name__ == '__main__':
    main()
