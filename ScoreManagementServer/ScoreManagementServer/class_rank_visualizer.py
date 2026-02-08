#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
班级排名趋势分析工具
查看整个班级每个学生的排名变化情况
"""

import sqlite3
from datetime import datetime, timedelta
import os
import sys

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# 数据库路径 - 使用相对路径
DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'StudentData.db')

# 科目映射
SUBJECT_MAPPING = {
    1: '语文', 2: '数学', 3: '英语',
    4: '物理', 5: '化学', 6: '生物',
    7: '政治', 8: '历史', 9: '地理',
    10: '总分'
}


def connect_db():
    """连接数据库"""
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def get_all_classes(conn):
    """获取所有班级"""
    cursor = conn.cursor()
    cursor.execute("""
        SELECT DISTINCT ClassName
        FROM Students
        WHERE ClassName IS NOT NULL AND ClassName != ''
        ORDER BY ClassName
    """)
    return cursor.fetchall()


def get_students_in_class(conn, class_name):
    """获取班级所有学生"""
    cursor = conn.cursor()
    cursor.execute("""
        SELECT StudentId, StudentNumber, StudentName
        FROM Students
        WHERE ClassName = ?
        ORDER BY StudentNumber
    """, (class_name,))
    return cursor.fetchall()


def get_class_rank_trend(conn, class_name, subject_id, start_date=None, end_date=None, rank_type='grade'):
    """获取班级某科目的排名趋势

    Args:
        conn: 数据库连接
        class_name: 班级名称
        subject_id: 科目ID
        start_date: 开始日期
        end_date: 结束日期
        rank_type: 排名类型，'class'为班级排名，'grade'为年级排名（默认）
    """
    rank_field = 's.ClassRank' if rank_type == 'class' else 's.GradeRank'

    sql = f"""
        SELECT
            e.ExamId,
            e.ExamName,
            e.ExamDate,
            s.StudentId,
            st.StudentNumber,
            st.StudentName,
            {rank_field} as Rank
        FROM Scores s
        JOIN Exams e ON s.ExamId = e.ExamId
        JOIN Students st ON s.StudentId = st.StudentId
        WHERE st.ClassName = ? AND s.SubjectId = ?
    """

    params = [class_name, subject_id]

    if start_date:
        sql += " AND e.ExamDate >= ?"
        params.append(start_date)

    if end_date:
        sql += " AND e.ExamDate <= ?"
        params.append(end_date)

    sql += " ORDER BY e.ExamDate, st.StudentNumber"

    cursor = conn.cursor()
    cursor.execute(sql, params)
    return cursor.fetchall()


def get_all_subjects_with_scores(conn, class_name):
    """获取班级有成绩记录的科目"""
    cursor = conn.cursor()
    cursor.execute("""
        SELECT DISTINCT
            sb.SubjectId,
            sb.SubjectName,
            sb.SortOrder,
            COUNT(s.ScoreId) as RecordCount
        FROM Subjects sb
        JOIN Scores s ON sb.SubjectId = s.SubjectId
        JOIN Students st ON s.StudentId = st.StudentId
        WHERE st.ClassName = ?
        GROUP BY sb.SubjectId, sb.SubjectName, sb.SortOrder
        ORDER BY sb.SortOrder
    """, (class_name,))
    return cursor.fetchall()


def group_exams_by_semester(exams):
    """按学期分组考试

    Args:
        exams: 考试列表，每个元素是 {'ExamId', 'ExamName', 'ExamDate'}

    Returns:
        list: 学期列表，每个元素是 {'name', 'start_date', 'end_date', 'exams'}
    """
    if not exams:
        return []

    semesters = []

    # 将考试按日期分组
    # 假设每半年为一个学期
    # 根据考试名称或日期推断学期
    for exam in exams:
        exam_date = exam['ExamDate']
        year = int(exam_date.split('-')[0])
        month = int(exam_date.split('-')[1])

        # 根据月份判断学期
        # 第一学期：9月-次年2月
        # 第二学期：3月-8月
        if month >= 9:
            semester_name = f"{year}年第一学期（上学期）"
        elif month <= 2:
            semester_name = f"{year-1}年第一学期（上学期）"
        elif 3 <= month <= 8:
            semester_name = f"{year}年第二学期（下学期）"
        else:
            semester_name = f"{year}年未知学期"

        # 查找是否已有该学期
        found = False
        for semester in semesters:
            if semester['name'] == semester_name:
                semester['exams'].append(exam)
                # 更新结束日期
                if exam_date > semester['end_date']:
                    semester['end_date'] = exam_date
                # 更新开始日期
                if exam_date < semester['start_date']:
                    semester['start_date'] = exam_date
                found = True
                break

        if not found:
            semesters.append({
                'name': semester_name,
                'start_date': exam_date,
                'end_date': exam_date,
                'exams': [exam]
            })

    # 按开始日期排序
    semesters.sort(key=lambda x: x['start_date'])

    return semesters


def sort_students_by_change(table_data):
    """按总变化排序学生

    排序逻辑：
    1. 总变化为正（进步）的排在前面，按变化值从大到小
    2. 总变化为0的排在中间
    3. 总变化为负（退步）的排在后面，按变化值从大到小（即退步从少到多）

    Args:
        table_data: 学生数据列表

    Returns:
        list: 排序后的学生数据列表
    """
    def sort_key(student):
        total_change = student['total_change']

        # 如果总变化为None，放到最后
        if total_change is None:
            return (-2, 0)

        # 总变化为正（进步）：使用正数排序，从大到小
        # 总变化为0：排序码为-1
        # 总变化为负（退步）：转换为绝对值，从小到大（退步从少到多）
        if total_change > 0:
            return (0, -total_change)  # 进步最多的在前
        elif total_change == 0:
            return (-1, 0)  # 保持不变的在中间
        else:
            return (1, abs(total_change))  # 退步最少的在前

    return sorted(table_data, key=sort_key)


def print_sorted_table(table_data, sorted_exams):
    """打印排序后的表格

    Args:
        table_data: 学生数据列表
        sorted_exams: 排序后的考试列表
    """
    # 构建表头
    header_parts = ['序号', '姓名', '学号']
    for exam_id, exam_info in sorted_exams:
        header_parts.append(f"{exam_info['ExamName'][:10]}")
        if sorted_exams.index((exam_id, exam_info)) < len(sorted_exams) - 1:
            header_parts.append('变化')
    header_parts.append('总变化')

    header = f"{'序号':^5} {'姓名':^8} {'学号':^13}"
    for exam_id, exam_info in sorted_exams:
        header += f" {exam_info['ExamName'][:10]:^12}"
        if sorted_exams.index((exam_id, exam_info)) < len(sorted_exams) - 1:
            header += f" {'变化':^8}"

    print(header)
    print(f"{'-'*150}")

    # 输出表格
    for idx, row_data in enumerate(table_data, 1):
        formatted_name = format_name(row_data['name'])
        number_str = row_data['number'] or ''

        # 构建每行的数据
        row_parts = [idx, formatted_name, number_str]

        exam_ranks = row_data['exam_ranks']
        changes = row_data['changes']

        # 获取每次考试的排名
        for i, rank in enumerate(exam_ranks):
            rank_str = str(rank) if rank is not None else 'N/A'
            row_parts.append(f"{rank_str:^12}")

            # 如果不是最后一次考试，添加变化列
            if i < len(exam_ranks) - 1:
                if i < len(changes):
                    change_type, change_value = changes[i]
                    if change_type == 'improve':
                        change_str = f"\033[31m↑{change_value}\033[0m"  # 红色
                    elif change_type == 'decline':
                        change_str = f"\033[32m↓{change_value}\033[0m"  # 绿色
                    elif change_type == 'same':
                        change_str = "→"
                    else:
                        change_str = "-"
                else:
                    change_str = "-"

                row_parts.append(f"{change_str:^8}")

        # 计算总变化（第一次 vs 最后一次）
        first_valid = next((r for r in exam_ranks if r is not None), None)
        last_valid = next((r for r in reversed(exam_ranks) if r is not None), None)
        if first_valid is not None and last_valid is not None:
            total_change = first_valid - last_valid
            if total_change > 0:
                total_change_str = f"\033[31m↑{total_change}\033[0m"  # 红色
            elif total_change < 0:
                total_change_str = f"\033[32m↓{abs(total_change)}\033[0m"  # 绿色
            else:
                total_change_str = "→"
        else:
            total_change_str = "-"

        row_parts.append(f"{total_change_str:^10}")

        # 输出行
        row_str = f"{row_parts[0]:^5} {row_parts[1]:^8} {row_parts[2]:^13}"
        for part in row_parts[3:]:
            row_str += f" {part}"
        print(row_str)

    print(f"{'-'*150}")


def format_name(name):
    """格式化姓名：两个字中间加空格"""
    if len(name) == 2:
        return f"{name[0]}  {name[1]}"
    return name


def print_class_rank_summary(scores, class_name, subject_name, decline_threshold=5, rank_type='grade'):
    """打印班级排名变化摘要

    Args:
        scores: 成绩数据
        class_name: 班级名称
        subject_name: 科目名称
        decline_threshold: 退步显示阈值
        rank_type: 排名类型，'class'为班级排名，'grade'为年级排名

    Returns:
        tuple: (improvements, declines, no_change) 用于后续生成Excel
    """
    if not scores:
        print(f"⚠️  {class_name}没有{subject_name}成绩记录")
        return [], [], []

    rank_type_name = "年级排名" if rank_type == 'grade' else "班级排名"

    # 按学生分组
    students_data = {}
    for s in scores:
        if s['StudentId'] not in students_data:
            students_data[s['StudentId']] = {
                'name': s['StudentName'],
                'number': s['StudentNumber'],
                'records': []
            }
        students_data[s['StudentId']]['records'].append(s)

    # 调试信息：显示有多少学生有记录
    print(f"\n📋 调试信息: 共获取 {len(scores)} 条成绩记录，涉及 {len(students_data)} 名学生")

    # 找出进步和退步最大的学生
    improvements = []
    declines = []
    no_change = []

    # 统计每个学生的考试次数
    exam_count_distribution = {}
    for student_id, data in students_data.items():
        records = sorted(data['records'], key=lambda x: datetime.strptime(x['ExamDate'], '%Y-%m-%d'))
        exam_count = len(records)
        exam_count_distribution[exam_count] = exam_count_distribution.get(exam_count, 0) + 1

    # 显示考试次数分布
    print(f"📊 考试次数分布:", end=" ")
    for exam_count in sorted(exam_count_distribution.keys()):
        print(f"{exam_count}次考试: {exam_count_distribution[exam_count]}人", end="; ")
    print()

    for student_id, data in students_data.items():
        records = sorted(data['records'], key=lambda x: datetime.strptime(x['ExamDate'], '%Y-%m-%d'))

        if len(records) >= 2:
            first_rank = records[0]['Rank']
            last_rank = records[-1]['Rank']

            # 只有当第一次和最后一次排名都存在时，才计算进步/退步
            if first_rank is not None and last_rank is not None:
                change = first_rank - last_rank  # 正数表示进步（排名上升）

                if change > 0:
                    improvements.append((change, data['name'], data['number'],
                                       records[0]['ExamName'], records[0]['Rank'],
                                       records[-1]['ExamName'], records[-1]['Rank']))
                elif change < 0:
                    declines.append((abs(change), data['name'], data['number'],
                                   records[0]['ExamName'], records[0]['Rank'],
                                   records[-1]['ExamName'], records[-1]['Rank']))
                else:
                    no_change.append((data['name'], data['number']))

    improvements.sort(reverse=True)
    declines.sort(reverse=True)

    print(f"\n{'='*102}")
    print(f"📊 {class_name} - {subject_name}{rank_type_name}变化分析")
    print(f"{'='*102}")

    # 显示所有进步的学生
    if improvements:
        print(f"\n📈 进步学生（共{len(improvements)}人）:")
        header = f"{'序号':^6} {'姓名':>9} {'学号':^13} {'变化':^8} {'初始考试':^20} {'初始排名':^8} {'最近考试':^20} {'最新排名':^8}"
        print(header)
        print(f"{'-'*102}")
        for i, (change, name, number, first_exam, first_rank, last_exam, last_rank) in enumerate(improvements, 1):
            formatted_name = format_name(name)
            first_rank_str = str(first_rank) if first_rank else 'N/A'
            last_rank_str = str(last_rank) if last_rank else 'N/A'
            number_str = number or ''
            print(f"{i:^6} {formatted_name:>9} {number_str:^13} +{change:^6} {first_exam:^20} {first_rank_str:^8} {last_exam:^20} {last_rank_str:^8}")

    # 只显示退步明显的学生（超过阈值）
    if declines:
        significant_declines = [d for d in declines if d[0] >= decline_threshold]
        if significant_declines:
            print(f"\n📉 明显退步学生（退步{decline_threshold}名及以上，共{len(significant_declines)}人）:")
            header = f"{'序号':^6} {'姓名':>9} {'学号':^13} {'变化':^8} {'初始考试':^20} {'初始排名':^8} {'最近考试':^20} {'最新排名':^8}"
            print(header)
            print(f"{'-'*102}")
            for i, (change, name, number, first_exam, first_rank, last_exam, last_rank) in enumerate(significant_declines, 1):
                formatted_name = format_name(name)
                first_rank_str = str(first_rank) if first_rank else 'N/A'
                last_rank_str = str(last_rank) if last_rank else 'N/A'
                number_str = number or ''
                print(f"{i:^6} {formatted_name:>9} {number_str:^13} -{change:^6} {first_exam:^20} {first_rank_str:^8} {last_exam:^20} {last_rank_str:^8}")
        else:
            print(f"\n📉 没有学生退步{decline_threshold}名及以上")

    # 显示统计信息
    total_students = len(students_data)
    valid_students = len(improvements) + len(declines) + len(no_change)
    incomplete_students = total_students - valid_students  # 排名缺失的学生
    print(f"\n📌 统计摘要:")
    print(f"  班级总人数: {total_students}人")
    print(f"  有效记录: {valid_students}人")
    if incomplete_students > 0:
        print(f"  排名缺失: {incomplete_students}人（某次考试排名为空）")
    print(f"  进步人数: {len(improvements)}人")
    print(f"  退步人数: {len(declines)}人")
    print(f"  保持不变: {len(no_change)}人")

    if improvements:
        avg_improvement = sum(i[0] for i in improvements) / len(improvements)
        print(f"  平均进步: {avg_improvement:.1f}名")

    if declines:
        avg_decline = sum(d[0] for d in declines) / len(declines)
        print(f"  平均退步: {avg_decline:.1f}名")

    print(f"{'='*102}")

    return improvements, declines, no_change


def print_multi_exam_rank_trend(scores, class_name, subject_name, rank_type='grade'):
    """打印多次考试的详细排名变化表格

    Args:
        scores: 成绩数据
        class_name: 班级名称
        subject_name: 科目名称
        rank_type: 排名类型，'class'为班级排名，'grade'为年级排名

    Returns:
        dict: 用于生成Excel的数据，包含每个学生在每次考试中的排名和变化
    """
    if not scores:
        print(f"⚠️  {class_name}没有{subject_name}成绩记录")
        return None

    rank_type_name = "年级排名" if rank_type == 'grade' else "班级排名"

    # 按学生分组
    students_data = {}
    for s in scores:
        if s['StudentId'] not in students_data:
            students_data[s['StudentId']] = {
                'name': s['StudentName'],
                'number': s['StudentNumber'],
                'records': []
            }
        students_data[s['StudentId']]['records'].append(s)

    # 获取所有考试列表并按日期排序
    all_exams = {}
    for s in scores:
        if s['ExamId'] not in all_exams:
            all_exams[s['ExamId']] = {
                'ExamName': s['ExamName'],
                'ExamDate': s['ExamDate']
            }

    # 按日期排序考试
    sorted_exams = sorted(all_exams.items(), key=lambda x: datetime.strptime(x[1]['ExamDate'], '%Y-%m-%d'))

    if len(sorted_exams) < 2:
        print(f"⚠️  该班级{subject_name}只有{len(sorted_exams)}次考试记录，无法显示多次考试变化")
        return None

    print(f"\n{'='*150}")
    print(f"📊 {class_name} - {subject_name}{rank_type_name}多次考试详细分析")
    print(f"{'='*150}")

    # 构建表格数据
    table_data = []
    for student_id, data in students_data.items():
        records = sorted(data['records'], key=lambda x: datetime.strptime(x['ExamDate'], '%Y-%m-%d'))

        row_data = {
            'student_id': student_id,
            'name': data['name'],
            'number': data['number'],
            'exam_ranks': {}  # {exam_id: rank}
        }

        # 记录每次考试的排名
        for record in records:
            row_data['exam_ranks'][record['ExamId']] = record['Rank']

        table_data.append(row_data)

    # 构建表头
    header_parts = ['序号', '姓名', '学号']
    for exam_id, exam_info in sorted_exams:
        header_parts.append(f"{exam_info['ExamName'][:10]}")
        if sorted_exams.index((exam_id, exam_info)) < len(sorted_exams) - 1:
            header_parts.append('变化')  # 添加变化列
    header_parts.append('总分变化')
    header = f"{'序号':^5} {'姓名':^8} {'学号':^13}"
    for exam_id, exam_info in sorted_exams:
        header += f" {exam_info['ExamName'][:10]:^12}"
        if sorted_exams.index((exam_id, exam_info)) < len(sorted_exams) - 1:
            header += f" {'变化':^8}"

    print(header)
    print(f"{'-'*150}")

    # 输出表格
    excel_data = []
    total_improvements = 0
    total_declines = 0

    for idx, row_data in enumerate(table_data, 1):
        formatted_name = format_name(row_data['name'])
        number_str = row_data['number'] or ''

        # 构建每行的数据
        row_parts = [idx, formatted_name, number_str]

        exam_ranks = []
        changes = []

        # 获取每次考试的排名
        for exam_id, exam_info in sorted_exams:
            rank = row_data['exam_ranks'].get(exam_id)
            exam_ranks.append(rank)

        # 先输出每次考试的排名
        for rank in exam_ranks:
            if rank is not None:
                rank_str = str(rank)
            else:
                rank_str = 'N/A'
            row_parts.append(f"{rank_str:^12}")

        # 计算相邻两次考试之间的变化
        for i in range(len(exam_ranks) - 1):
            current_rank = exam_ranks[i]
            next_rank = exam_ranks[i + 1]
            if current_rank is not None and next_rank is not None:
                change = current_rank - next_rank  # 正数表示进步(排名上升)
                if change > 0:
                    change_str = f"\033[31m↑{change}\033[0m"  # 红色表示进步
                    changes.append(('improve', change))
                    total_improvements += 1
                elif change < 0:
                    change_str = f"\033[32m↓{abs(change)}\033[0m"  # 绿色表示退步
                    changes.append(('decline', abs(change)))
                    total_declines += 1
                else:
                    change_str = "→"
                    changes.append(('same', 0))
            else:
                change_str = "-"
                changes.append(('invalid', 0))

            row_parts.append(f"{change_str:^8}")

            # 计算总变化（第一次 vs 最后一次）
            first_valid = next((r for r in exam_ranks if r is not None), None)
            last_valid = next((r for r in reversed(exam_ranks) if r is not None), None)
            if first_valid is not None and last_valid is not None:
                total_change = first_valid - last_valid
                if total_change > 0:
                    total_change_str = f"\033[31m↑{total_change}\033[0m"  # 红色表示进步
                elif total_change < 0:
                    total_change_str = f"\033[32m↓{abs(total_change)}\033[0m"  # 绿色表示退步
                else:
                    total_change_str = "→"
            else:
                total_change_str = "-"

        row_parts.append(f"{total_change_str:^10}")

        # 输出行
        row_str = f"{row_parts[0]:^5} {row_parts[1]:^8} {row_parts[2]:^13}"
        for part in row_parts[3:]:
            row_str += f" {part}"
        print(row_str)

        # 保存用于Excel的数据
        excel_row = {
            'name': row_data['name'],
            'number': row_data['number'],
            'exam_ranks': exam_ranks,
            'changes': changes,
            'total_change': total_change if first_valid is not None and last_valid is not None else None
        }
        excel_data.append(excel_row)

    # 统计信息
    print(f"\n📌 统计摘要:")
    print(f"  班级人数: {len(table_data)}人")
    print(f"  考试次数: {len(sorted_exams)}次")
    print(f"  进步人次: {total_improvements}次")
    print(f"  退步人次: {total_declines}次")

    print(f"{'='*150}")

    return {
        'sorted_exams': sorted_exams,
        'table_data': excel_data,
        'total_improvements': total_improvements,
        'total_declines': total_declines
    }


def print_class_rank_sum_trend(scores, class_name, subject_name):
    """打印班级学科排名总和变化

    Args:
        scores: 成绩数据
        class_name: 班级名称
        subject_name: 科目名称
    """
    if not scores:
        print(f"⚠️  {class_name}没有{subject_name}成绩记录")
        return

    # 按考试分组，计算每次考试的所有学生排名总和
    exam_rank_sums = {}
    exam_students_count = {}

    for s in scores:
        exam_name = s['ExamName']
        exam_date = s['ExamDate']
        rank = s['Rank']

        if exam_name not in exam_rank_sums:
            exam_rank_sums[exam_name] = {
                'ExamDate': exam_date,
                'RankSum': 0,
                'StudentCount': 0
            }

        if rank is not None:
            exam_rank_sums[exam_name]['RankSum'] += rank
            exam_rank_sums[exam_name]['StudentCount'] += 1

    if not exam_rank_sums:
        print(f"⚠️  没有有效的排名数据")
        return

    # 按日期排序
    sorted_exams = sorted(exam_rank_sums.items(), key=lambda x: datetime.strptime(x[1]['ExamDate'], '%Y-%m-%d'))

    if len(sorted_exams) < 2:
        print(f"⚠️  该班级{subject_name}只有{len(sorted_exams)}次考试记录，无法计算排名总和变化")
        return

    # 计算排名总和变化
    rank_sum_changes = []
    for i in range(1, len(sorted_exams)):
        prev_exam_name, prev_exam_data = sorted_exams[i-1]
        curr_exam_name, curr_exam_data = sorted_exams[i]

        prev_rank_sum = prev_exam_data['RankSum']
        curr_rank_sum = curr_exam_data['RankSum']
        prev_student_count = prev_exam_data['StudentCount']
        curr_student_count = curr_exam_data['StudentCount']

        change = prev_rank_sum - curr_rank_sum  # 正数表示进步（排名总和变小）

        if change > 0:
            trend = "📈 进步"
            trend_icon = "↑"
        elif change < 0:
            trend = "📉 倒退"
            trend_icon = "↓"
        else:
            trend = "➡️ 持平"
            trend_icon = "→"

        rank_sum_changes.append({
            'prev_exam': prev_exam_name,
            'curr_exam': curr_exam_name,
            'prev_date': prev_exam_data['ExamDate'],
            'curr_date': curr_exam_data['ExamDate'],
            'prev_rank_sum': prev_rank_sum,
            'curr_rank_sum': curr_rank_sum,
            'change': change,
            'trend': trend,
            'trend_icon': trend_icon,
            'prev_count': prev_student_count,
            'curr_count': curr_student_count
        })

    print(f"\n{'='*100}")
    print(f"📊 {class_name} - {subject_name}年级排名总和变化分析")
    print(f"{'='*100}")

    # 显示变化表格
    print(f"\n{'序号':^6} {'上次考试':^18} {'上次排名总和':^12} {'当前考试':^18} {'当前排名总和':^12} {'变化':^10} {'趋势'}")
    print(f"{'-'*100}")

    for i, change_data in enumerate(rank_sum_changes, 1):
        prev_exam = change_data['prev_exam']
        curr_exam = change_data['curr_exam']
        prev_rank_sum = change_data['prev_rank_sum']
        curr_rank_sum = change_data['curr_rank_sum']
        change = change_data['change']
        trend = change_data['trend']
        trend_icon = change_data['trend_icon']
        prev_count = change_data['prev_count']
        curr_count = change_data['curr_count']

        change_str = f"+{change}" if change > 0 else f"{change}"

        print(f"{i:^6} {prev_exam:^18} {prev_rank_sum:^12} {curr_exam:^18} {curr_rank_sum:^12} {change_str:^10} {trend}")

    # 统计摘要
    total_improvements = sum(1 for c in rank_sum_changes if c['change'] > 0)
    total_declines = sum(1 for c in rank_sum_changes if c['change'] < 0)
    total_no_change = sum(1 for c in rank_sum_changes if c['change'] == 0)

    total_change = sum(c['change'] for c in rank_sum_changes)
    avg_change = total_change / len(rank_sum_changes) if rank_sum_changes else 0

    print(f"\n📌 统计摘要:")
    print(f"  考试次数: {len(sorted_exams)}次")
    print(f"  进步次数: {total_improvements}次")
    print(f"  倒退次数: {total_declines}次")
    print(f"  持平次数: {total_no_change}次")
    print(f"  总体变化: {total_change:+.0f}")
    print(f"  平均变化: {avg_change:+.1f}")

    if total_change > 0:
        print(f"  ✅ 总体趋势：进步")
    elif total_change < 0:
        print(f"  ⚠️  总体趋势：倒退")
    else:
        print(f"  ➡️ 总体趋势：持平")

    print(f"{'='*100}")


def generate_multi_exam_excel(data, class_name, subject_name, rank_type_name):
    """生成多次考试的Excel报告

    Args:
        data: 多次考试数据
        class_name: 班级名称
        subject_name: 科目名称
        rank_type_name: 排名类型名称
    """
    if not HAS_OPENPYXL:
        print("❌ 缺少openpyxl库，无法生成Excel文件")
        print("请运行: pip install openpyxl")
        return False

    try:
        # 创建工作簿
        wb = Workbook()

        # 删除默认的sheet
        wb.remove(wb.active)

        # 定义样式
        header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        cell_alignment = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 红色字体样式（表示进步）
        improve_font = Font(name='微软雅黑', size=11, color='C00000')
        # 绿色字体样式（表示退步）
        decline_font = Font(name='微软雅黑', size=11, color='00B050')
        # 灰色字体样式（表示无效）
        gray_font = Font(name='微软雅黑', size=11, color='999999')

        # 创建详细排名sheet
        ws = wb.create_sheet(title='详细排名变化')

        # 构建表头
        sorted_exams = data['sorted_exams']
        headers = ['序号', '姓名', '学号']
        for exam_id, exam_info in sorted_exams:
            headers.append(exam_info['ExamName'])
            if sorted_exams.index((exam_id, exam_info)) < len(sorted_exams) - 1:
                headers.append('变化')
        headers.append('总变化')

        # 写入表头
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # 写入数据
        for idx, excel_row in enumerate(data['table_data'], 2):
            row_values = [idx - 1, excel_row['name'], excel_row['number']]

            # 写入每次考试的排名
            exam_ranks = excel_row['exam_ranks']
            changes = excel_row['changes']

            for i, rank in enumerate(exam_ranks):
                if rank is not None:
                    row_values.append(rank)
                else:
                    row_values.append('N/A')

                # 如果不是最后一次考试，添加变化列
                if i < len(exam_ranks) - 1:
                    if i < len(changes):
                        change_type, change_value = changes[i]
                        if change_type == 'improve':
                            row_values.append(f"↑{change_value}")
                        elif change_type == 'decline':
                            row_values.append(f"↓{change_value}")
                        elif change_type == 'same':
                            row_values.append('→')
                        else:
                            row_values.append('-')
                    else:
                        row_values.append('-')

            # 添加总变化
            total_change = excel_row['total_change']
            if total_change is not None:
                if total_change > 0:
                    row_values.append(f"↑{total_change}")
                elif total_change < 0:
                    row_values.append(f"↓{abs(total_change)}")
                else:
                    row_values.append('→')
            else:
                row_values.append('-')

            # 写入单元格并应用样式
            for col, value in enumerate(row_values, 1):
                cell = ws.cell(row=idx, column=col, value=value)
                cell.alignment = cell_alignment
                cell.border = thin_border

                # 对变化列和总变化列应用颜色
                col_header = ws.cell(row=1, column=col).value
                if col_header == '变化' or col_header == '总变化':
                    if isinstance(value, str) and '↑' in value:
                        cell.font = improve_font
                    elif isinstance(value, str) and '↓' in value:
                        cell.font = decline_font

        # 调整列宽
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 15
        for col in range(4, ws.max_column + 1):
            ws.column_dimensions[chr(64 + col)].width = 12

        # 冻结首行和前三列
        ws.freeze_panes = 'D2'

        # 解析班级信息生成文件名
        semester = ""
        class_short = class_name

        # 提取学期信息（上学期/下学期）
        if len(sorted_exams) > 0:
            first_exam_date = sorted_exams[0][1]['ExamDate']
            year = int(first_exam_date.split('-')[0])
            month = int(first_exam_date.split('-')[1])

            if month >= 9 or month <= 2:
                semester = "上学期"
            elif 3 <= month <= 8:
                semester = "下学期"

        # 提取班级名称（如107班）
        if "班" in class_name:
            for part in class_name.split():
                if "班" in part:
                    class_short = part
                    break

        # 生成文件名：班级+上/下学期+科目
        if semester:
            filename = f"{class_short}{semester}{subject_name}.xlsx"
        else:
            filename = f"{class_short}{subject_name}.xlsx"

        # 保存到桌面
        desktop = os.path.join(os.path.expanduser('~'), 'Desktop')
        file_path = os.path.join(desktop, filename)

        wb.save(file_path)
        print(f"\n✅ Excel报告已生成: {file_path}")
        print(f"  - 学生人数: {len(data['table_data'])}人")
        print(f"  - 考试次数: {len(sorted_exams)}次")

        return True

    except Exception as e:
        print(f"\n❌ 生成Excel失败: {e}")
        import traceback
        traceback.print_exc()
        return False


def generate_excel_report(improvements, declines, class_name, first_exam, last_exam, subject_name):
    """生成Excel报告

    Args:
        improvements: 进步学生列表
        declines: 退步学生列表
        class_name: 班级名称
        first_exam: 开始考试名称
        last_exam: 结束考试名称
        subject_name: 科目名称
    """
    if not HAS_OPENPYXL:
        print("❌ 缺少openpyxl库，无法生成Excel文件")
        print("请运行: pip install openpyxl")
        return False

    try:
        # 创建工作簿
        wb = Workbook()

        # 删除默认的sheet
        wb.remove(wb.active)

        # 定义样式
        header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        cell_alignment = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 创建进步sheet
        if improvements:
            ws_improve = wb.create_sheet(title='进步学生')
            # 写入表头
            headers = ['序号', '姓名', '学号', '进步名次', '初始考试', '初始排名', '最近考试', '最新排名']
            for col, header in enumerate(headers, 1):
                cell = ws_improve.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

            # 写入数据
            for idx, student in enumerate(improvements, 2):
                # 支持元组格式：(change, name, number, first_exam, first_rank, last_exam, last_rank)
                if isinstance(student, tuple):
                    change, name, number, first_exam, first_rank, last_exam, last_rank = student
                    data = [idx - 1, name, number, change, first_exam, first_rank, last_exam, last_rank]
                else:
                    # 字典格式
                    data = [
                        idx - 1,
                        student['name'],
                        student['number'],
                        student['change'],
                        student['first_exam'],
                        student['first_rank'],
                        student['last_exam'],
                        student['last_rank']
                    ]
                for col, value in enumerate(data, 1):
                    cell = ws_improve.cell(row=idx, column=col, value=value)
                    cell.alignment = cell_alignment
                    cell.border = thin_border

            # 调整列宽
            ws_improve.column_dimensions['A'].width = 6
            ws_improve.column_dimensions['B'].width = 10
            ws_improve.column_dimensions['C'].width = 15
            ws_improve.column_dimensions['D'].width = 10
            ws_improve.column_dimensions['E'].width = 20
            ws_improve.column_dimensions['F'].width = 10
            ws_improve.column_dimensions['G'].width = 20
            ws_improve.column_dimensions['H'].width = 10

            # 冻结首行
            ws_improve.freeze_panes = 'A2'
        else:
            ws_improve = wb.create_sheet(title='进步学生')
            ws_improve.cell(row=1, column=1, value='无进步学生记录')

        # 创建退退sheet
        if declines:
            ws_decline = wb.create_sheet(title='退步学生')
            # 写入表头
            headers = ['序号', '姓名', '学号', '退步名次', '初始考试', '初始排名', '最近考试', '最新排名']
            for col, header in enumerate(headers, 1):
                cell = ws_decline.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
                cell.alignment = header_alignment
                cell.border = thin_border

            # 写入数据
            for idx, student in enumerate(declines, 2):
                # 支持元组格式：(change, name, number, first_exam, first_rank, last_exam, last_rank)
                if isinstance(student, tuple):
                    change, name, number, first_exam, first_rank, last_exam, last_rank = student
                    data = [idx - 1, name, number, change, first_exam, first_rank, last_exam, last_rank]
                else:
                    # 字典格式
                    data = [
                        idx - 1,
                        student['name'],
                        student['number'],
                        student['change'],
                        student['first_exam'],
                        student['first_rank'],
                        student['last_exam'],
                        student['last_rank']
                    ]
                for col, value in enumerate(data, 1):
                    cell = ws_decline.cell(row=idx, column=col, value=value)
                    cell.alignment = cell_alignment
                    cell.border = thin_border

            # 调整列宽
            ws_decline.column_dimensions['A'].width = 6
            ws_decline.column_dimensions['B'].width = 10
            ws_decline.column_dimensions['C'].width = 15
            ws_decline.column_dimensions['D'].width = 10
            ws_decline.column_dimensions['E'].width = 20
            ws_decline.column_dimensions['F'].width = 10
            ws_decline.column_dimensions['G'].width = 20
            ws_decline.column_dimensions['H'].width = 10

            # 冻结首行
            ws_decline.freeze_panes = 'A2'
        else:
            ws_decline = wb.create_sheet(title='退步学生')
            ws_decline.cell(row=1, column=1, value='无退步学生记录')

        # 解析班级信息生成文件名
        semester = ""
        class_short = class_name

        # 使用当前日期推断学期
        import datetime as dt
        current_date = dt.datetime.now()
        month = current_date.month

        if month >= 9 or month <= 2:
            semester = "上学期"
        elif 3 <= month <= 8:
            semester = "下学期"

        # 提取班级名称（如107班）
        if "班" in class_name:
            for part in class_name.split():
                if "班" in part:
                    class_short = part
                    break

        # 生成文件名：班级+上/下学期+科目
        if semester:
            filename = f"{class_short}{semester}{subject_name}.xlsx"
        else:
            filename = f"{class_short}{subject_name}.xlsx"

        # 保存到桌面
        desktop = os.path.join(os.path.expanduser('~'), 'Desktop')
        file_path = os.path.join(desktop, filename)

        wb.save(file_path)
        print(f"\n✅ Excel报告已生成: {file_path}")
        print(f"  - 进步学生: {len(improvements)}人")
        print(f"  - 退步学生: {len(declines)}人")

        return True

    except Exception as e:
        print(f"\n❌ 生成Excel失败: {e}")
        import traceback
        traceback.print_exc()
        return False




def main():
    print("=" * 100)
    print("                              班级排名趋势分析工具")
    print("=" * 100)

    conn = connect_db()

    try:
        while True:  # 主循环：支持连续查询
            # 第一步：选择班级
            classes = get_all_classes(conn)

            if not classes:
                print("❌ 数据库中没有班级信息")
                return

            print(f"\n找到 {len(classes)} 个班级:")
            for i, c in enumerate(classes, 1):
                print(f"  {i}. {c['ClassName']}")

            choice = input("\n请选择班级编号: ").strip()
            if not choice.isdigit() or int(choice) < 1 or int(choice) > len(classes):
                print("❌ 无效选择")
                continue

            class_name = classes[int(choice) - 1]['ClassName']

            # 第二步：显示可用科目
            subjects = get_all_subjects_with_scores(conn, class_name)

            if not subjects:
                print(f"⚠️  {class_name}没有成绩记录")
                continue

            print(f"\n{'='*100}")
            print(f"  {class_name} 的考试科目")
            print(f"{'='*100}")

            for i, (subject_id, subject_name, sort_order, count) in enumerate(subjects, 1):
                print(f"  {i}. {subject_name} ({count}条记录)")

            choice = input(f"\n请选择科目（1-{len(subjects)}）: ").strip()
            if not choice.isdigit() or int(choice) < 1 or int(choice) > len(subjects):
                print("❌ 无效选择")
                continue

            subject_id, subject_name = subjects[int(choice) - 1][:2]

            # 第三步：选择分析类型
            print(f"\n{'='*100}")
            print(f"  选择分析类型")
            print(f"{'='*100}")
            print(f"  1. 学生个人排名变化分析（对比最近两次考试）")
            print(f"  2. 班级学科排名总和变化分析")
            print(f"  3. 多次考试详细排名变化分析（按学期查看所有考试）")
            print(f"{'='*100}")

            analysis_type_choice = input(f"\n请选择分析类型（1-3，默认1）: ").strip()
            if not analysis_type_choice or not analysis_type_choice.isdigit():
                analysis_type = 1
            else:
                analysis_type = int(analysis_type_choice)

            if analysis_type == 2:
                # 班级学科排名总和变化分析
                scores = get_class_rank_trend(conn, class_name, subject_id, None, None, 'grade')
                print_class_rank_sum_trend(scores, class_name, subject_name)
                continue

            if analysis_type == 1:
                # 选项1：学生个人排名变化分析（对比最近两次考试）
                # 第四步：选择排名类型
                print(f"\n{'='*100}")
                print(f"  选择排名类型")
                print(f"{'='*100}")
                print(f"  1. 年级排名（默认）")
                print(f"  2. 班级排名")

                rank_type_choice = input(f"\n请选择排名类型（1-2，默认1）: ").strip()
                if not rank_type_choice or not rank_type_choice.isdigit():
                    rank_type = 'grade'
                elif rank_type_choice == '1':
                    rank_type = 'grade'
                else:
                    rank_type = 'class'

                rank_type_name = "年级排名" if rank_type == 'grade' else "班级排名"
                print(f"  排名类型: {rank_type_name}")

                # 获取最近两次考试
                cursor = conn.cursor()
                cursor.execute("""
                    SELECT DISTINCT e.ExamId, e.ExamDate, e.ExamName
                    FROM Exams e
                    JOIN Scores s ON e.ExamId = s.ExamId
                    JOIN Students st ON s.StudentId = st.StudentId
                    WHERE s.SubjectId = ? AND st.ClassName = ?
                    ORDER BY e.ExamDate DESC
                    LIMIT 2
                """, (subject_id, class_name))
                recent_exams = cursor.fetchall()

                if len(recent_exams) < 2:
                    print(f"⚠️  该科目只有{len(recent_exams)}次考试，无法对比最近两次考试")
                    continue

                # 获取这两次考试的日期范围
                exam_dates = [e['ExamDate'] for e in recent_exams]
                start_date = min(exam_dates)
                end_date = max(exam_dates)

                print(f"\n时间范围: 最近两次考试")
                print(f"  考试1: {recent_exams[1]['ExamName']} ({recent_exams[1]['ExamDate']})")
                print(f"  考试2: {recent_exams[0]['ExamName']} ({recent_exams[0]['ExamDate']})")

                # 第五步：选择退步阈值
                decline_threshold = input("\n退步显示阈值（退步多少名以上才显示，默认5名）: ").strip()
                if decline_threshold and decline_threshold.isdigit():
                    decline_threshold = int(decline_threshold)
                else:
                    decline_threshold = 5
                print(f"退步阈值: {decline_threshold}名")

                # 第六步：获取数据并分析
                print(f"\n正在获取{class_name}的{subject_name}{rank_type_name}数据...")
                scores = get_class_rank_trend(conn, class_name, subject_id, start_date, end_date, rank_type)

                if not scores:
                    print(f"⚠️  该班级没有{subject_name}成绩记录")
                    continue

                # 打印摘要并获取数据
                improvements, declines, no_change = print_class_rank_summary(scores, class_name, subject_name, decline_threshold, rank_type)

                # 获取第一次和最后一次考试名称
                sorted_scores = sorted(scores, key=lambda x: datetime.strptime(x['ExamDate'], '%Y-%m-%d'))
                first_exam = sorted_scores[0]['ExamName'] if sorted_scores else None
                last_exam = sorted_scores[-1]['ExamName'] if sorted_scores else None

                # 询问是否生成Excel
                if improvements or declines:
                    while True:
                        export_choice = input(f"\n是否生成Excel报告？(y/n): ").strip().lower()
                        if export_choice in ['y', 'n']:
                            break
                        print("❌ 请输入 y 或 n")

                    if export_choice == 'y':
                        generate_excel_report(improvements, declines, class_name, first_exam, last_exam, subject_name)

                continue

            if analysis_type == 3:
                # 分析类型3：多次考试详细排名变化分析
                # 选择排名类型
                print(f"\n{'='*100}")
                print(f"  选择排名类型")
                print(f"{'='*100}")
                print(f"  1. 年级排名（默认）")
                print(f"  2. 班级排名")

                rank_type_choice = input(f"\n请选择排名类型（1-2，默认1）: ").strip()
                if not rank_type_choice or not rank_type_choice.isdigit():
                    rank_type = 'grade'
                elif rank_type_choice == '1':
                    rank_type = 'grade'
                else:
                    rank_type = 'class'

                rank_type_name = "年级排名" if rank_type == 'grade' else "班级排名"
                print(f"  排名类型: {rank_type_name}")

                # 获取该科目所有考试
                cursor = conn.cursor()
                cursor.execute("""
                    SELECT DISTINCT e.ExamId, e.ExamName, e.ExamDate
                    FROM Exams e
                    JOIN Scores s ON e.ExamId = s.ExamId
                    WHERE s.SubjectId = ?
                    ORDER BY e.ExamDate
                """, (subject_id,))
                all_exams = cursor.fetchall()

                if len(all_exams) < 2:
                    print(f"⚠️  该科目只有{len(all_exams)}次考试，无法分析多次考试变化")
                    continue

                # 按学期分组考试
                semesters = group_exams_by_semester(all_exams)

                if not semesters:
                    print(f"⚠️  没有找到可以分析的学期")
                    continue

                # 选择学期
                print(f"\n{'='*100}")
                print(f"  选择学期")
                print(f"{'='*100}")
                for i, semester in enumerate(semesters, 1):
                    exam_count = len(semester['exams'])
                    start_date = semester['start_date']
                    end_date = semester['end_date']
                    print(f"  {i}. {semester['name']} ({start_date} 至 {end_date}, 共{exam_count}次考试)")

                semester_choice = input(f"\n请选择学期（1-{len(semesters)}）: ").strip()
                if not semester_choice or not semester_choice.isdigit() or int(semester_choice) < 1 or int(semester_choice) > len(semesters):
                    print("❌ 无效选择")
                    continue

                selected_semester = semesters[int(semester_choice) - 1]
                print(f"\n已选择: {selected_semester['name']}")

                # 获取该学期的考试日期范围
                start_date = selected_semester['start_date']
                end_date = selected_semester['end_date']

                # 获取数据
                print(f"\n正在获取{class_name}的{subject_name}{rank_type_name}数据...")
                scores = get_class_rank_trend(conn, class_name, subject_id, start_date, end_date, rank_type)

                if not scores:
                    print(f"⚠️  该班级在这个学期没有{subject_name}成绩记录")
                    continue

                # 打印多次考试详细排名变化
                data = print_multi_exam_rank_trend(scores, class_name, subject_name, rank_type)

                if data:
                    # 询问排序方式
                    print(f"\n{'='*100}")
                    print(f"  选择排序方式")
                    print(f"{'='*100}")
                    print(f"  1. 按总变化排序（进步最多的在前）")
                    print(f"  2. 按学号排序（默认）")
                    print(f"{'='*100}")

                    sort_choice = input(f"\n请选择排序方式（1-2，默认2）: ").strip()
                    if not sort_choice or not sort_choice.isdigit():
                        sort_choice = '2'

                    if sort_choice == '1':
                        # 按总变化排序
                        data['table_data'] = sort_students_by_change(data['table_data'])

                        # 重新打印排序后的表格
                        print(f"\n{'='*150}")
                        print(f"📊 {class_name} - {subject_name}{rank_type_name}多次考试详细分析（按总变化排序）")
                        print(f"{'='*150}")
                        print_sorted_table(data['table_data'], data['sorted_exams'])

                    # 询问是否生成Excel
                    while True:
                        export_choice = input(f"\n是否生成Excel报告？(y/n): ").strip().lower()
                        if export_choice in ['y', 'n']:
                            break
                        print("❌ 请输入 y 或 n")

                    if export_choice == 'y':
                        generate_multi_exam_excel(data, class_name, subject_name, rank_type_name)

                continue

            # 询问是否继续
            while True:
                continue_query = input("\n是否继续查询？(y/n): ").strip().lower()
                if continue_query in ['y', 'n']:
                    break
                print("❌ 请输入 y 或 n")

            if continue_query == 'n':
                print("\n感谢使用班级排名趋势分析工具，再见！")
                break  # 退出主循环

    except Exception as e:
        print(f"\n❌ 发生错误: {e}")
        import traceback
        traceback.print_exc()

    finally:
        conn.close()


if __name__ == '__main__':
    main()
