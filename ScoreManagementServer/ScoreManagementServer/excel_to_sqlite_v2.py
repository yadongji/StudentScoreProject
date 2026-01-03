#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel导入工具 V2.0 - 简化解耦版
支持查询和成绩趋势分析
"""

import sqlite3
import os
import sys
import re
from datetime import datetime

try:
    from openpyxl import load_workbook
except ImportError:
    print("❌ 缺少openpyxl库")
    print("请运行: pip install openpyxl")
    sys.exit(1)

# 配置
DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "StudentData.db")
SCHEMA_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "database_schema_simple.sql")

# 科目映射表 - 使用科目名作为key，数据库SubjectId作为value
SUBJECT_IDS = {
    '语文': 1,
    '数学': 2,
    '英语': 3,
    '物理': 4,
    '化学': 5,
    '生物': 6,
    '政治': 7,
    '历史': 8,
    '地理': 9,
    '总分': 10  # SubjectId为10时表示总分
}

print("=" * 50)
print("   高中成绩管理系统 V2.0 - Excel导入工具")
print("=" * 50)
print()


def create_database():
    """创建数据库"""
    print(f"\n📝 正在创建数据库...")

    if not os.path.exists(SCHEMA_FILE):
        print(f"❌ 数据库架构文件不存在: {SCHEMA_FILE}")
        return False

    with open(SCHEMA_FILE, 'r', encoding='utf-8') as f:
        sql = f.read()

    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()

    # 分割并执行SQL语句
    statements = sql.split(';')
    for stmt in statements:
        stmt = stmt.strip()
        if stmt and not stmt.startswith('--'):
            try:
                cursor.execute(stmt)
            except Exception as e:
                if "already exists" not in str(e).lower():
                    print(f"⚠️  {str(e)[:60]}")

    conn.commit()
    conn.close()

    print(f"✅ 数据库创建成功: {DB_PATH}")
    return True


def update_database_schema():
    """更新数据库架构：允许StudentNumber为NULL"""
    print(f"\n🔄 正在检查并更新数据库架构...")

    try:
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()

        # 检查 Students 表是否存在
        cursor.execute("""
            SELECT name FROM sqlite_master
            WHERE type='table' AND name='Students'
        """)
        table_exists = cursor.fetchone()

        if not table_exists:
            print("⚠️  Students 表不存在，跳过架构更新")
            conn.close()
            return False

        # 检查 StudentNumber 是否允许 NULL
        cursor.execute("PRAGMA table_info(Students)")
        columns = cursor.fetchall()

        studentnumber_notnull = False
        for col in columns:
            col_name = col[1]  # 列名在索引1
            notnull = col[3]  # notnull 约束在索引3
            if col_name == 'StudentNumber' and notnull == 1:
                studentnumber_notnull = True
                break

        if studentnumber_notnull:
            print("🔧 检测到 StudentNumber 字段不允许为空，正在更新架构...")

            # SQLite 不支持直接修改列约束，需要重建表
            # 1. 创建新表
            cursor.execute("""
                CREATE TABLE Students_new (
                    StudentId INTEGER PRIMARY KEY AUTOINCREMENT,
                    StudentNumber TEXT UNIQUE,
                    StudentName TEXT NOT NULL,
                    ClassName TEXT,
                    Gender TEXT CHECK(Gender IN ('男', '女')),
                    EnrollmentDate TEXT,
                    IsActive INTEGER DEFAULT 1,
                    CreatedAt TEXT DEFAULT (datetime('now', 'localtime')),
                    UpdatedAt TEXT DEFAULT (datetime('now', 'localtime'))
                )
            """)

            # 2. 复制数据
            cursor.execute("""
                INSERT INTO Students_new (StudentId, StudentNumber, StudentName, ClassName, Gender, EnrollmentDate, IsActive, CreatedAt, UpdatedAt)
                SELECT StudentId, StudentNumber, StudentName, ClassName, Gender, EnrollmentDate, IsActive, CreatedAt, UpdatedAt
                FROM Students
            """)

            # 3. 删除旧表
            cursor.execute("DROP TABLE Students")

            # 4. 重命名新表
            cursor.execute("ALTER TABLE Students_new RENAME TO Students")

            # 5. 重建索引
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_students_number ON Students(StudentNumber)")
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_students_name ON Students(StudentName)")
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_students_class ON Students(ClassName)")

            conn.commit()
            print("✅ 数据库架构更新成功：StudentNumber 字段现在可以为空")
        else:
            print("✅ 数据库架构已是最新版本")

        conn.close()
        return True

    except Exception as e:
        print(f"❌ 数据库架构更新失败: {e}")
        import traceback
        traceback.print_exc()
        return False


def import_students():
    """导入学生信息(仅学号和姓名)"""
    print("\n📋 导入学生信息")
    print("-" * 50)

    file_path = input("请输入Excel文件路径 (如: 107学生考号(新).xlsx): ").strip()

    if not os.path.exists(file_path):
        print("❌ 文件不存在!")
        return

    default_class = input("请输入默认班级名称 (可选,直接回车跳过): ").strip()

    print("\n⏳ 正在导入...")

    try:
        wb = load_workbook(filename=file_path, read_only=True)
        ws = wb.active

        # 读取表头
        headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]

        # 创建字段映射
        col_map = {}
        for idx, header in enumerate(headers):
            if header:
                if '学号' in header or '考号' in header:
                    col_map['学号'] = idx
                elif '姓名' in header:
                    col_map['姓名'] = idx
                elif '班级' in header:
                    col_map['班级'] = idx
                elif '性别' in header:
                    col_map['性别'] = idx

        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()

        success = 0
        updated = 0
        failed = 0

        # 从第2行开始读取数据
        for row in ws.iter_rows(min_row=2):
            # 读取学号
            student_number_cell = row[col_map.get('学号', 0)] if col_map.get('学号') is not None else None
            student_number = str(student_number_cell.value).strip() if student_number_cell and student_number_cell.value else ""

            # 读取姓名
            student_name_cell = row[col_map.get('姓名', 1)] if col_map.get('姓名') is not None else None
            student_name = str(student_name_cell.value).strip() if student_name_cell and student_name_cell.value else ""

            # 读取班级
            class_name_cell = row[col_map.get('班级', 2)] if col_map.get('班级') is not None else None
            class_name = str(class_name_cell.value).strip() if class_name_cell and class_name_cell.value else ""

            # 使用默认班级或从Excel读取
            if not class_name and default_class:
                class_name = default_class

            if not student_number or student_number == "None" or not student_name:
                continue

            try:
                # 检查是否已存在
                cursor.execute("SELECT StudentId FROM Students WHERE StudentNumber = ?", (student_number,))
                existing = cursor.fetchone()

                if existing:
                    # 更新
                    cursor.execute("""
                        UPDATE Students SET
                            StudentName = ?,
                            ClassName = ?,
                            UpdatedAt = datetime('now', 'localtime')
                        WHERE StudentNumber = ?
                    """, (student_name, class_name, student_number))
                    updated += 1
                else:
                    # 插入新学生
                    cursor.execute("""
                        INSERT INTO Students (StudentNumber, StudentName, ClassName)
                        VALUES (?, ?, ?)
                    """, (student_number, student_name, class_name))
                    success += 1
            except Exception as e:
                failed += 1
                print(f"❌ 学号 {student_number}: {e}")

        conn.commit()
        conn.close()
        wb.close()

        print(f"\n✅ 导入完成: 新增 {success} 条, 更新 {updated} 条, 失败 {failed} 条")

    except Exception as e:
        print(f"\n❌ 导入失败: {e}")
        import traceback
        traceback.print_exc()


def find_matching_class(cursor, input_class):
    """根据输入的班级名，在数据库中查找匹配的班级

    匹配规则：
    1. 精确匹配
    2. 包含匹配（输入班级名包含在数据库班级名中）
    3. 反向包含匹配（数据库班级名包含在输入班级名中）
    4. 数字匹配（提取输入班级名中的数字，与数据库班级名中的数字匹配）
    """
    if not input_class or input_class == "None":
        return None

    # 获取所有已有班级
    cursor.execute("SELECT DISTINCT ClassName FROM Students WHERE ClassName IS NOT NULL AND ClassName != '' ORDER BY ClassName")
    all_classes = [row[0] for row in cursor.fetchall()]

    if not all_classes:
        return input_class  # 数据库中没有班级，直接使用输入的

    # 1. 精确匹配
    if input_class in all_classes:
        return input_class

    # 2. 包含匹配
    for db_class in all_classes:
        if input_class in db_class:
            return db_class
        if db_class in input_class:
            return db_class

    # 3. 数字匹配
    import re
    # 提取输入班级名中的数字
    input_numbers = re.findall(r'\d+', input_class)
    if input_numbers:
        input_num = input_numbers[0]
        for db_class in all_classes:
            db_numbers = re.findall(r'\d+', db_class)
            if db_numbers and db_numbers[0] == input_num:
                return db_class

    # 都不匹配，使用输入的班级名
    return input_class


def update_students_info():
    """更新学生信息（根据姓名查找并更新学号等信息）"""
    print("\n📋 更新学生信息")
    print("-" * 50)
    print("说明：此功能根据姓名在数据库中查找学生，并更新学号等信息")
    print("      如果数据库中不存在该学生，将自动添加新学生")
    print("      新增学生时，班级名称会自动匹配数据库中的标准班级名")
    print("      Excel表格格式与导入学生信息相同")

    file_path = input("请输入Excel文件路径 (如: 107学生考号(新).xlsx): ").strip()

    if not os.path.exists(file_path):
        print("❌ 文件不存在!")
        return

    print("\n⏳ 正在处理...")

    try:
        wb = load_workbook(filename=file_path, read_only=True)
        ws = wb.active

        # 读取表头
        headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]

        # 创建字段映射
        col_map = {}
        for idx, header in enumerate(headers):
            if header:
                if '学号' in header or '考号' in header:
                    col_map['学号'] = idx
                elif '姓名' in header:
                    col_map['姓名'] = idx
                elif '班级' in header:
                    col_map['班级'] = idx
                elif '性别' in header:
                    col_map['性别'] = idx

        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()

        updated = 0
        inserted = 0
        skipped = 0
        failed = 0

        # 从第2行开始读取数据
        for row in ws.iter_rows(min_row=2):
            # 读取姓名（必须）
            student_name_cell = row[col_map.get('姓名', 1)] if col_map.get('姓名') is not None else None
            student_name = str(student_name_cell.value).strip() if student_name_cell and student_name_cell.value else ""

            # 读取学号
            student_number_cell = row[col_map.get('学号', 0)] if col_map.get('学号') is not None else None
            student_number = str(student_number_cell.value).strip() if student_number_cell and student_number_cell.value else ""

            # 读取班级
            class_name_cell = row[col_map.get('班级', 2)] if col_map.get('班级') is not None else None
            input_class_name = str(class_name_cell.value).strip() if class_name_cell and class_name_cell.value else ""

            if not student_name or student_name == "None":
                continue

            if not student_number or student_number == "None":
                print(f"⚠️  跳过: {student_name} | 学号为空")
                skipped += 1
                continue

            # 如果有班级信息，匹配数据库中的标准班级名
            matched_class = None
            if input_class_name:
                matched_class = find_matching_class(cursor, input_class_name)

            try:
                # 根据姓名查找学生
                cursor.execute("SELECT StudentId, StudentNumber, ClassName FROM Students WHERE StudentName = ?", (student_name,))
                existing = cursor.fetchone()

                if existing:
                    # 学生已存在，更新
                    student_id, old_number, old_class = existing
                    has_change = False

                    # 检查学号是否需要更新
                    # 条件：Excel中学号不为空，且（数据库中学号为空 或 学号不同）
                    old_number_empty = not old_number or old_number.strip() == ""
                    if student_number and (old_number_empty or student_number != old_number):
                        # 检查新学号是否已被其他学生使用
                        cursor.execute("SELECT StudentId, StudentName FROM Students WHERE StudentNumber = ? AND StudentId != ?", (student_number, student_id))
                        conflict_record = cursor.fetchone()
                        if conflict_record:
                            # 存在学号冲突，废弃旧学号，以新学号为准
                            conflict_id, conflict_name = conflict_record
                            cursor.execute("""
                                UPDATE Students SET
                                    StudentNumber = NULL,
                                    UpdatedAt = datetime('now', 'localtime')
                                WHERE StudentId = ?
                            """, (conflict_id,))
                            print(f"⚠️  废弃学号: {conflict_name} | 学号 {student_number} 已被清空")

                        # 更新当前学生学号
                        cursor.execute("""
                            UPDATE Students SET
                                StudentNumber = ?,
                                UpdatedAt = datetime('now', 'localtime')
                            WHERE StudentId = ?
                        """, (student_number, student_id))
                        if old_number_empty:
                            print(f"✅ 补充: {student_name} | 学号: 空 → {student_number}")
                        else:
                            print(f"✅ 更新: {student_name} | 学号: {old_number} → {student_number}")
                        has_change = True

                    # 检查班级是否需要更新
                    # 条件：Excel中班级不为空，且（数据库中班级为空 或 班级不同）
                    old_class_empty = not old_class or old_class.strip() == ""
                    if matched_class and (old_class_empty or matched_class != old_class):
                        cursor.execute("""
                            UPDATE Students SET
                                ClassName = ?,
                                UpdatedAt = datetime('now', 'localtime')
                            WHERE StudentId = ?
                        """, (matched_class, student_id))
                        if old_class_empty:
                            print(f"✅ 补充: {student_name} | 班级: 空 → {matched_class} (自动匹配: {input_class_name})")
                        elif input_class_name and input_class_name != matched_class:
                            print(f"✅ 更新: {student_name} | 班级: {old_class or '空'} → {matched_class} (自动匹配: {input_class_name})")
                        else:
                            print(f"✅ 更新: {student_name} | 班级: {old_class or '空'} → {matched_class}")
                        has_change = True

                    if has_change:
                        updated += 1
                    else:
                        print(f"⏭️  跳过: {student_name} | 信息无变化")
                        skipped += 1

                else:
                    # 学生不存在，插入新学生
                    # 检查学号是否已存在
                    cursor.execute("SELECT StudentId, StudentName FROM Students WHERE StudentNumber = ?", (student_number,))
                    number_record = cursor.fetchone()
                    if number_record:
                        # 存在学号冲突，废弃旧学号
                        number_id, number_name = number_record
                        cursor.execute("""
                            UPDATE Students SET
                                StudentNumber = NULL,
                                UpdatedAt = datetime('now', 'localtime')
                            WHERE StudentId = ?
                        """, (number_id,))
                        print(f"⚠️  废弃学号: {number_name} | 学号 {student_number} 已被清空")

                    # 插入新学生
                    display_class = matched_class if matched_class else "未设置"
                    if input_class_name and matched_class and input_class_name != matched_class:
                        print(f"➕ 新增: {student_name} | 学号: {student_number} | 班级: {matched_class} (自动匹配: {input_class_name})")
                    else:
                        print(f"➕ 新增: {student_name} | 学号: {student_number} | 班级: {display_class}")
                    cursor.execute("""
                        INSERT INTO Students (StudentNumber, StudentName, ClassName)
                        VALUES (?, ?, ?)
                    """, (student_number, student_name, matched_class))
                    inserted += 1

            except Exception as e:
                failed += 1
                print(f"❌ 处理失败 {student_name}: {e}")

        conn.commit()
        conn.close()
        wb.close()

        print(f"\n✅ 处理完成:")
        print(f"  新增学生: {inserted} 条")
        print(f"  更新学生: {updated} 条")
        print(f"  跳过: {skipped} 条")
        print(f"  失败: {failed} 条")

    except Exception as e:
        print(f"\n❌ 更新失败: {e}")
        import traceback
        traceback.print_exc()


def detect_sheet_columns(headers):
    """
    智能检测Excel表的列位置
    优先级：
    1. 检测学号列（包含"学号"或"考号"二字）
    2. 如果没有学号，检测姓名列（包含"姓名"二字）
    3. 检测学科成绩列（学科名或学科名+成绩/分数）
    4. 检测排名列（成绩列附近查找"班级排名"/"班次"/"年级排名"）
    """
    # 列索引映射（从0开始）
    col_map = {}

    # 1. 优先检测学号列
    for idx, header in enumerate(headers):
        if header and ('学号' in header or '考号' in header):
            col_map['学号'] = idx
            break

    # 2. 如果没找到学号，检测姓名列
    if '学号' not in col_map:
        for idx, header in enumerate(headers):
            if header and '姓名' in header:
                col_map['姓名'] = idx
                break

    # 3. 检测班级列
    for idx, header in enumerate(headers):
        if header and '班级' in header:
            col_map['班级'] = idx
            break

    # 4. 检测总分相关列
    for idx, header in enumerate(headers):
        if header:
            if '总分分数' in header or '总分' == header:
                col_map['总分_score'] = idx
            elif '总分校名次' in header or '总分班级排名' in header:
                col_map['总分_grade_rank'] = idx
            elif '总分班名次' in header or '总分班级名次' in header:
                col_map['总分_class_rank'] = idx

    # 5. 检测各学科成绩和排名列
    for subject_name in ['语文', '数学', '英语', '物理', '化学', '生物', '政治', '历史', '地理']:
        # 成绩列：支持 "学科名", "学科名成绩", "学科名分数"
        score_col = None
        for idx, header in enumerate(headers):
            if header:
                if (header == subject_name or
                    f'{subject_name}成绩' in header or
                    f'{subject_name}分数' in header):
                    col_map[f'{subject_name}_score'] = idx
                    score_col = idx
                    break

        # 班级排名列：成绩列后查找 "班名次", "班级名次", "班级排名", "班次"
        if score_col is not None:
            # 优先查找带学科前缀的排名列
            class_rank_col = None
            for idx in range(score_col + 1, min(score_col + 3, len(headers))):
                header = headers[idx]
                if header and (f'{subject_name}班名次' in header or
                              f'{subject_name}班级名次' in header or
                              f'{subject_name}班级排名' in header):
                    col_map[f'{subject_name}_class_rank'] = idx
                    class_rank_col = idx
                    break

            # 如果没找到带学科前缀的，查找通用的"班次"、"班级排名"等
            if class_rank_col is None:
                for idx in range(score_col + 1, min(score_col + 3, len(headers))):
                    header = headers[idx]
                    if header and ('班次' in header or '班级排名' in header or '班级名次' in header):
                        col_map[f'{subject_name}_class_rank'] = idx
                        break

        # 年级/学校排名列：成绩列后查找 "年级名次", "年级排名", "校名次", "校次"
        if score_col is not None:
            # 优先查找带学科前缀的排名列
            for idx in range(score_col + 1, min(score_col + 5, len(headers))):
                header = headers[idx]
                if header and (f'{subject_name}年级名次' in header or
                              f'{subject_name}年级排名' in header or
                              f'{subject_name}校名次' in header or
                              f'{subject_name}校次' in header):
                    col_map[f'{subject_name}_grade_rank'] = idx
                    break

            # 如果没找到带学科前缀的，查找通用的"年级排名"、"年级名次"、"校名次"、"校次"等
            # 注意：需要判断是否已被其他学科占用
            found = False
            for idx in range(score_col + 1, min(score_col + 5, len(headers))):
                header = headers[idx]
                if header and ('年级排名' in header or '年级名次' in header or '校名次' in header or '校次' in header):
                    # 检查该列是否已被分配给其他学科
                    is_assigned = any(f'{subj}_grade_rank' in col_map and col_map[f'{subj}_grade_rank'] == idx
                                    for subj in ['语文', '数学', '英语', '物理', '化学', '生物', '政治', '历史', '地理'])
                    if not is_assigned:
                        col_map[f'{subject_name}_grade_rank'] = idx
                        found = True
                        break
                elif found:
                    break

    return col_map


def import_scores():
    """导入学生成绩"""
    print("\n📊 导入学生成绩")
    print("-" * 50)

    file_path = input("请输入Excel文件路径 (如: 107班物化生成绩.xlsx): ").strip()

    if not os.path.exists(file_path):
        print("❌ 文件不存在!")
        return

    exam_id = input("请输入考试ID: ").strip()
    if not exam_id.isdigit():
        print("❌ 无效的考试ID!")
        return
    exam_id = int(exam_id)

    print("\n⏳ 正在导入...")

    try:
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()

        # 检查考试是否存在
        cursor.execute("SELECT * FROM Exams WHERE ExamId = ?", (exam_id,))
        if not cursor.fetchone():
            print("❌ 考试不存在,请先创建考试!")
            conn.close()
            return

        wb = load_workbook(filename=file_path, read_only=True)
        ws = wb.active

        # 读取表头
        headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]

        print(f"\n📋 Excel文件信息:")
        print(f"  总行数: {ws.max_row}")
        print(f"  总列数: {ws.max_column}")
        print(f"  表头列: {headers}")

        # 使用智能列检测
        col_map = detect_sheet_columns(headers)

        print(f"\n🔍 字段映射:")
        print(f"  {col_map}")

        # 显示匹配模式
        if '学号' in col_map:
            print(f"\n✅ 使用学号匹配模式")
        elif '姓名' in col_map:
            print(f"\n✅ 使用姓名匹配模式 (根据姓名查找学号)")
        else:
            print(f"\n⚠️  警告: Excel中既没有学号列也没有姓名列!")

        # 检查总分科目是否存在，不存在则创建（SubjectId固定为10）
        cursor.execute("SELECT SubjectId FROM Subjects WHERE SubjectId = ?", (10,))
        total_subject = cursor.fetchone()
        if not total_subject:
            print("\n⚠️  数据库中不存在'总分'科目（SubjectId=10），正在添加...")
            cursor.execute("INSERT INTO Subjects (SubjectId, SubjectName, SubjectCode) VALUES (10, '总分', 'TOTAL')")
            conn.commit()
            print(f"✅ 已添加'总分'科目 (SubjectId: 10)")
        else:
            # 检查科目名称是否正确
            cursor.execute("SELECT SubjectName FROM Subjects WHERE SubjectId = ?", (10,))
            subject_name = cursor.fetchone()
            if subject_name and subject_name[0] != '总分':
                print(f"⚠️  SubjectId=10的科目名称是'{subject_name[0]}'，正在更新为'总分'...")
                cursor.execute("UPDATE Subjects SET SubjectName = '总分' WHERE SubjectId = ?", (10,))
                conn.commit()
                print(f"✅ 已更新为'总分'")

        success = 0
        failed = 0
        processed = 0
        has_student_number = '学号' in col_map

        print(f"\n开始处理数据...")

        for row in ws.iter_rows(min_row=2):
            processed += 1

            student_id = None
            student_number = ""
            student_name = ""
            class_name = ""

            # 读取班级（如果有）
            if '班级' in col_map:
                class_cell = row[col_map['班级']]
                class_name = str(class_cell.value).strip() if class_cell and class_cell.value else ""

            # 优先使用学号，如果没有则使用姓名查找
            if has_student_number:
                # 读取学号
                student_number_cell = row[col_map.get('学号', 0)]
                student_number = str(student_number_cell.value).strip() if student_number_cell and student_number_cell.value else ""

                if student_number and student_number != "None":
                    # 根据学号查找
                    cursor.execute("SELECT StudentId, StudentName FROM Students WHERE StudentNumber = ?", (student_number,))
                    student = cursor.fetchone()
                    if student:
                        student_id, student_name = student
                    else:
                        print(f"⚠️  第{processed}行: 学号 '{student_number}' 不存在,跳过")
                        failed += 1
                        continue
                else:
                    print(f"⚠️  第{processed}行: 学号为空,跳过")
                    failed += 1
                    continue
            else:
                # 没有学号列，使用姓名查找
                if '姓名' in col_map:
                    name_cell = row[col_map['姓名']]
                    student_name = str(name_cell.value).strip() if name_cell and name_cell.value else ""

                    if student_name and student_name != "None":
                        # 根据姓名查找（可以加上班级筛选，如果有班级信息）
                        if class_name:
                            # 如果有班级信息，优先匹配同班级的学生
                            cursor.execute("SELECT StudentId, StudentNumber FROM Students WHERE StudentName = ? AND ClassName = ?", (student_name, class_name))
                            students = cursor.fetchall()
                            if len(students) == 0:
                                # 同班级没找到，去掉班级限制再找
                                cursor.execute("SELECT StudentId, StudentNumber FROM Students WHERE StudentName = ?", (student_name,))
                                students = cursor.fetchall()
                        else:
                            cursor.execute("SELECT StudentId, StudentNumber FROM Students WHERE StudentName = ?", (student_name,))
                            students = cursor.fetchall()

                        if len(students) == 0:
                            print(f"⚠️  第{processed}行: 未找到姓名为 '{student_name}' 的学生" + (f"(班级:{class_name})" if class_name else "") + ",跳过")
                            failed += 1
                            continue
                        elif len(students) > 1:
                            print(f"⚠️  第{processed}行: 姓名为 '{student_name}' 的学生有{len(students)}个,使用第一个")
                            student_id, student_number = students[0]
                        else:
                            student_id, student_number = students[0]
                    else:
                        print(f"⚠️  第{processed}行: 姓名为空,跳过")
                        failed += 1
                        continue
                else:
                    print(f"⚠️  第{processed}行: Excel中既没有学号列也没有姓名列,跳过")
                    failed += 1
                    continue

            # 如果还是没找到学生ID，跳过
            if not student_id:
                if has_student_number:
                    print(f"⚠️  第{processed}行: 学号 '{student_number}' 未找到,跳过")
                else:
                    print(f"⚠️  第{processed}行: 姓名为 '{student_name}' 的学生未找到,跳过")
                failed += 1
                continue

            # 读取各科成绩（包括总分，SubjectId=10）
            for subject_name, subject_id in SUBJECT_IDS.items():
                # 只处理Excel中有对应列的科目
                if f'{subject_name}_score' not in col_map:
                    continue

                # 读取成绩
                score_cell = row[col_map.get(f'{subject_name}_score')]
                score = None
                if score_cell and score_cell.value:
                    try:
                        score = float(score_cell.value)
                    except:
                        pass

                # 读取班级排名
                class_rank = None
                class_rank_col = col_map.get(f'{subject_name}_class_rank')
                if class_rank_col is not None:
                    class_rank_cell = row[class_rank_col]
                    if class_rank_cell and class_rank_cell.value:
                        try:
                            class_rank = int(float(class_rank_cell.value))
                        except:
                            pass

                # 读取年级排名
                grade_rank = None
                grade_rank_col = col_map.get(f'{subject_name}_grade_rank')
                if grade_rank_col is not None:
                    grade_rank_cell = row[grade_rank_col]
                    if grade_rank_cell and grade_rank_cell.value:
                        try:
                            grade_rank = int(float(grade_rank_cell.value))
                        except:
                            pass

                # 如果有成绩则插入或更新
                if score is not None:
                    try:
                        # 检查成绩是否已存在
                        cursor.execute("""
                            SELECT ScoreId FROM Scores
                            WHERE ExamId = ? AND StudentId = ? AND SubjectId = ?
                        """, (exam_id, student_id, subject_id))
                        existing = cursor.fetchone()

                        if existing:
                            # 更新
                            cursor.execute("""
                                UPDATE Scores SET
                                    Score = ?,
                                    ClassRank = ?,
                                    GradeRank = ?,
                                    UpdatedAt = datetime('now', 'localtime')
                                WHERE ExamId = ? AND StudentId = ? AND SubjectId = ?
                            """, (score, class_rank, grade_rank, exam_id, student_id, subject_id))
                        else:
                            # 插入
                            cursor.execute("""
                                INSERT INTO Scores (
                                    ExamId, StudentId, SubjectId, Score, ClassRank, GradeRank
                                ) VALUES (?, ?, ?, ?, ?, ?)
                            """, (exam_id, student_id, subject_id, score, class_rank, grade_rank))

                        success += 1
                    except Exception as e:
                        failed += 1
                        student_info = f"学号{student_number}" if has_student_number else f"姓名'{student_name}'"
                        print(f"❌ {student_info} 科目 {subject_name}: {e}")

        conn.commit()
        conn.close()
        wb.close()

        print(f"\n========================================")
        print(f"导入结果:")
        print(f"  总行数: {processed}")
        print(f"  成功: {success}")
        print(f"  失败: {failed}")
        print(f"========================================")
        print(f"✅ 导入完成: 成功 {success} 条, 失败 {failed} 条")

    except Exception as e:
        print(f"\n❌ 导入失败: {e}")
        import traceback
        traceback.print_exc()


def create_exam():
    """创建考试"""
    print("\n创建新考试")
    print("-" * 50)

    exam_name = input("考试名称 (如: 2024年秋季期中考试): ").strip()
    exam_type = input("考试类型 (月考/期中考/期末考/模拟考/联考): ").strip()
    exam_date = input("考试日期 (如: 2024-11-15): ").strip()
    grade_name = input("年级 (高一/高二/高三): ").strip()
    term = input("学期 (上学期/下学期,可选): ").strip()
    academic_year = input("学年 (如: 2024-2025,可选): ").strip()

    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()

    cursor.execute("""
        INSERT INTO Exams (
            ExamName, ExamType, ExamDate, GradeName, Term, AcademicYear, IsPublished
        ) VALUES (?, ?, ?, ?, ?, ?, 0)
    """, (exam_name, exam_type, exam_date, grade_name, term, academic_year))

    exam_id = cursor.lastrowid
    conn.commit()
    conn.close()

    print(f"\n✅ 考试创建成功!")
    print(f"考试ID: {exam_id}")
    print(f"请记住这个ID,导入成绩时需要使用!")


def query_scores():
    """查询成绩"""
    print("\n🔍 查询成绩")
    print("-" * 50)

    # 查询方式
    print("查询方式:")
    print("1. 按学号查询")
    print("2. 按姓名查询")
    print("3. 查看所有学生成绩")
    choice = input("请选择 (1/2/3): ").strip()

    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()

    if choice == '1':
        # 按学号查询
        student_number = input("请输入学号: ").strip()
        query_student_by_number(cursor, student_number)
    elif choice == '2':
        # 按姓名查询
        student_name = input("请输入姓名: ").strip()
        query_student_by_name(cursor, student_name)
    elif choice == '3':
        # 查看所有
        query_all_students(cursor)
    else:
        print("❌ 无效选项!")

    conn.close()


def query_student_by_number(cursor, student_number):
    """按学号查询学生成绩及趋势"""
    # 查询学生信息
    cursor.execute("SELECT * FROM Students WHERE StudentNumber = ?", (student_number,))
    student = cursor.fetchone()

    if not student:
        print(f"❌ 未找到学号为 {student_number} 的学生!")
        return

    student_id, number, name, class_name, _, _, _ = student

    print(f"\n👤 学生信息")
    print(f"  学号: {number}")
    print(f"  姓名: {name}")
    print(f"  班级: {class_name if class_name else '未设置'}")

    # 查询各次考试成绩
    cursor.execute("""
        SELECT
            e.ExamName, e.ExamDate, e.ExamType,
            sb.SubjectName, s.Score, s.ClassRank, s.GradeRank
        FROM Scores s
        JOIN Exams e ON s.ExamId = e.ExamId
        JOIN Subjects sb ON s.SubjectId = sb.SubjectId
        WHERE s.StudentId = ?
        ORDER BY e.ExamDate DESC, sb.SortOrder
    """, (student_id,))

    scores = cursor.fetchall()

    if not scores:
        print("\n⚠️  该学生暂无成绩记录!")
        return

    print(f"\n📊 成绩记录 (共 {len(scores)} 条)")
    print("-" * 80)

    # 按考试分组
    from collections import defaultdict
    exam_scores = defaultdict(list)
    for score in scores:
        exam_name = score[0]
        exam_scores[exam_name].append(score)

    for exam_name, exam_data in exam_scores.items():
        print(f"\n【{exam_name}】")

        # 显示各科成绩
        for score in exam_data:
            _, exam_date, exam_type, subject, score_value, class_rank, grade_rank = score
            rank_info = f" (班排:{class_rank}/年排:{grade_rank})" if class_rank or grade_rank else ""
            print(f"  {subject}: {score_value}{rank_info}")

    # 查询趋势分析
    query_trend(cursor, student_id, name)


def query_student_by_name(cursor, student_name):
    """按姓名查询"""
    cursor.execute("SELECT * FROM Students WHERE StudentName = ?", (student_name,))
    students = cursor.fetchall()

    if not students:
        print(f"❌ 未找到姓名为 {student_name} 的学生!")
        return

    if len(students) == 1:
        # 只有一个学生,直接查询成绩
        student = students[0]
        query_student_by_number(cursor, student[1])
    else:
        # 多个学生,显示列表让用户选择
        print(f"\n找到 {len(students)} 个同名学生:")
        for idx, student in enumerate(students, 1):
            print(f"  {idx}. 学号: {student[1]}, 班级: {student[3] if student[3] else '未设置'}")

        choice = input("\n请选择序号: ").strip()
        if choice.isdigit() and 1 <= int(choice) <= len(students):
            selected = students[int(choice) - 1]
            query_student_by_number(cursor, selected[1])


def query_all_students(cursor):
    """查看所有学生成绩"""
    cursor.execute("""
        SELECT
            st.StudentNumber, st.StudentName, st.ClassName,
            e.ExamName, e.ExamDate,
            sb.SubjectName, s.Score
        FROM Scores s
        JOIN Exams e ON s.ExamId = e.ExamId
        JOIN Students st ON s.StudentId = st.StudentId
        JOIN Subjects sb ON s.SubjectId = sb.SubjectId
        ORDER BY e.ExamDate DESC, st.ClassName, st.StudentNumber
    """)

    scores = cursor.fetchall()

    if not scores:
        print("⚠️  暂无成绩记录!")
        return

    print(f"\n📊 所有学生成绩 (共 {len(scores)} 条)")
    print("-" * 80)
    print(f"{'学号':<12} {'姓名':<8} {'班级':<10} {'考试':<20} {'科目':<6} {'成绩':<6}")
    print("-" * 80)

    for score in scores[:50]:  # 只显示前50条
        number, name, class_name, exam_name, exam_date, subject, score_value = score
        class_str = class_name if class_name else '未设置'
        print(f"{number:<12} {name:<8} {class_str:<10} {exam_name:<20} {subject:<6} {score_value:<6}")

    if len(scores) > 50:
        print(f"\n... 还有 {len(scores) - 50} 条记录")


def query_trend(cursor, student_id, student_name):
    """查询成绩趋势分析"""
    print(f"\n📈 成绩趋势分析")
    print("-" * 80)

    cursor.execute("""
        SELECT
            sb.SubjectName,
            e.ExamId, e.ExamName, e.ExamDate,
            s.Score,
            s.ClassRank, s.GradeRank,
            -- 上次考试成绩
            (
                SELECT s2.Score
                FROM Scores s2
                JOIN Exams e2 ON s2.ExamId = e2.ExamId
                WHERE s2.StudentId = ? AND s2.SubjectId = s.SubjectId
                  AND e2.ExamDate < e.ExamDate
                ORDER BY e2.ExamDate DESC
                LIMIT 1
            ) as PrevScore,
            -- 上次考试排名
            (
                SELECT s2.ClassRank
                FROM Scores s2
                JOIN Exams e2 ON s2.ExamId = e2.ExamId
                WHERE s2.StudentId = ? AND s2.SubjectId = s.SubjectId
                  AND e2.ExamDate < e.ExamDate
                ORDER BY e2.ExamDate DESC
                LIMIT 1
            ) as PrevClassRank
        FROM Scores s
        JOIN Exams e ON s.ExamId = e.ExamId
        JOIN Subjects sb ON s.SubjectId = sb.SubjectId
        WHERE s.StudentId = ?
        ORDER BY sb.SortOrder, e.ExamDate DESC
    """, (student_id, student_id, student_id))

    trends = cursor.fetchall()

    # 按科目分组
    subject_trends = {}
    for trend in trends:
        subject = trend[0]
        if subject not in subject_trends:
            subject_trends[subject] = []
        subject_trends[subject].append(trend)

    # 显示趋势
    for subject, subject_data in subject_trends.items():
        print(f"\n【{subject}】")
        print(f"{'考试名称':<20} {'考试日期':<12} {'成绩':<6} {'班排':<5} {'上次成绩':<8} {'变化':<8} {'趋势':<6}")
        print("-" * 80)

        for trend in subject_data:
            exam_name, exam_id, exam_date, score, class_rank, grade_rank, prev_score, prev_class_rank = trend

            # 计算变化
            score_change = ""
            trend_mark = ""
            if prev_score:
                change = score - prev_score
                if change > 0:
                    score_change = f"+{change}"
                    trend_mark = "↑ 进步"
                elif change < 0:
                    score_change = f"{change}"
                    trend_mark = "↓ 退步"
                else:
                    score_change = "0"
                    trend_mark = "- 持平"

            prev_score_str = f"{prev_score}" if prev_score else "-"
            print(f"{exam_name:<20} {exam_date:<12} {score:<6} {class_rank if class_rank else '-':<5} {prev_score_str:<8} {score_change:<8} {trend_mark:<6}")


def show_statistics():
    """显示统计信息"""
    print("\n📈 数据库统计")
    print("-" * 50)

    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()

    # 统计数据
    cursor.execute("SELECT COUNT(*) FROM Students")
    student_count = cursor.fetchone()[0]

    cursor.execute("SELECT COUNT(*) FROM Exams")
    exam_count = cursor.fetchone()[0]

    cursor.execute("SELECT COUNT(*) FROM Scores")
    score_count = cursor.fetchone()[0]

    cursor.execute("SELECT COUNT(DISTINCT StudentId) FROM Scores")
    scored_student_count = cursor.fetchone()[0]

    cursor.execute("SELECT COUNT(DISTINCT SubjectId) FROM Scores")
    subject_count = cursor.fetchone()[0]

    print(f"学生总数: {student_count}")
    print(f"有成绩学生数: {scored_student_count}")
    print(f"考试数量: {exam_count}")
    print(f"成绩记录数: {score_count}")
    print(f"涉及科目数: {subject_count}")

    # 显示考试列表
    cursor.execute("SELECT ExamId, ExamName, ExamType, ExamDate, GradeName FROM Exams ORDER BY ExamDate DESC")
    exams = cursor.fetchall()

    if exams:
        print("\n📅 考试列表:")
        for exam in exams:
            print(f"  ID: {exam[0]}, {exam[1]} ({exam[2]}) - {exam[3]} - {exam[4]}")

    conn.close()


def main():
    """主函数"""
    # 检查数据库
    if not os.path.exists(DB_PATH):
        print(f"⚠️  数据库文件不存在: {DB_PATH}")
        choice = input("是否创建新数据库? (y/n): ").strip().lower()
        if choice == 'y':
            if not create_database():
                return
        else:
            return
    else:
        # 更新数据库架构
        update_database_schema()

    # 主菜单
    while True:
        print("\n" + "=" * 50)
        print("📋 主菜单")
        print("=" * 50)
        print("1. 导入学生信息")
        print("2. 导入学生成绩")
        print("3. 创建考试")
        print("4. 更新学生信息")
        print("5. 查询成绩")
        print("6. 查看数据库统计")
        print("7. 退出")
        print("=" * 50)
        choice = input("请输入选项 (1-7): ").strip()

        if choice == '1':
            import_students()
        elif choice == '2':
            import_scores()
        elif choice == '3':
            create_exam()
        elif choice == '4':
            update_students_info()
        elif choice == '5':
            query_scores()
        elif choice == '6':
            show_statistics()
        elif choice == '7':
            print("\n👋 感谢使用,再见!")
            break
        else:
            print("❌ 无效选项,请重新选择!")


if __name__ == "__main__":
    main()
