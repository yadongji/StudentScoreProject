#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
限时练成绩导入工具
支持导入限时练考试成绩，识别多sheet结构
"""

import openpyxl
import sqlite3
import re
from datetime import datetime

# 数据库路径
DB_PATH = r'E:\StudentScoreProject\ScoreManagementServer\ScoreManagementServer/StudentData.db'


def connect_db():
    """连接数据库"""
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def parse_exam_name(filename):
    """从文件名解析考试名称和科目"""
    # 文件名格式: 3.高一一级部物理限时练12.16-成绩排名-学生明细.xlsx
    # 提取: 物理限时练12.16
    match = re.search(r'高一.+?部(.+?)限时练([\d.]+)', filename)
    if match:
        subject = match.group(1).strip()
        date_str = match.group(2).strip()
        # 尝试解析日期 12.16 -> 2024-12-16
        try:
            if '.' in date_str:
                parts = date_str.split('.')
                month = int(parts[0])
                day = int(parts[1])
                exam_date = f"2024-{month:02d}-{day:02d}"
            else:
                exam_date = date_str
        except Exception as e:
            print(f"日期解析警告: {e}")
            exam_date = datetime.now().strftime('%Y-%m-%d')

        exam_name = f"{subject}限时练{date_str}"
        return exam_name, subject, exam_date

    return None, None, None


def find_subject_id(conn, subject_name):
    """查找科目ID"""
    cursor = conn.cursor()
    cursor.execute("SELECT SubjectId FROM Subjects WHERE SubjectName = ?", (subject_name,))
    result = cursor.fetchone()
    if result:
        return result['SubjectId']

    # 如果科目不存在，返回None
    return None


def create_time_limit_exam_if_not_exists(conn, exam_name, subject_name, subject_id, exam_date, grade_name='高一'):
    """创建限时练考试（如果不存在）"""
    cursor = conn.cursor()
    cursor.execute("SELECT ExamId FROM TimeLimitExams WHERE ExamName = ?", (exam_name,))
    result = cursor.fetchone()

    if result:
        return result['ExamId']

    # 确定学期
    try:
        month = int(exam_date.split('-')[1])
    except:
        month = 1
    term = '上学期' if month <= 7 else '下学期'

    # 确定学年
    try:
        year = int(exam_date.split('-')[0])
    except:
        year = 2024
    academic_year = f"{year}-{year+1}"

    # 创建新限时练考试
    cursor.execute("""
        INSERT INTO TimeLimitExams (ExamName, ExamDate, SubjectName, SubjectId, GradeName, Term, AcademicYear, Description)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
    """, (exam_name, exam_date, subject_name, subject_id, grade_name, term, academic_year, f"限时练考试"))
    conn.commit()
    return cursor.lastrowid


def clean_student_name(name):
    """清理学生姓名（去除后面的数字）"""
    if not name:
        return None

    # 去除后面的数字（如：张三1 -> 张三）
    name = str(name).strip()
    name = re.sub(r'\d+$', '', name)
    return name if name else None


def import_time_limit_sheet(conn, ws, exam_id, subject_id, class_name):
    """导入一个sheet的数据"""
    success_count = 0
    fail_count = 0
    errors = []

    try:
        # Excel结构说明：
        # 第1行：标题（忽略）
        # 第2,3行：合并的列名
        #   - 第2,3行合并第2列：学生姓名
        #   - 第2,3行合并第3列：学校学号
        #   - 第2行第4-8列合并：科目
        #   - 第3行第4列：成绩
        #   - 第3行第5列：班级排名
        #   - 第3行第7列：年级排名
        # 从第4行开始是数据

        # 直接按列位置读取
        col_name = 2          # 学生姓名
        col_school_number = 3    # 学校学号
        col_score = 4          # 成绩
        col_class_rank = 5       # 班级排名
        col_grade_rank = 7       # 年级排名

        print(f"\n  列位置: 姓名={col_name}, 学号={col_school_number}, 成绩={col_score}, 班级排名={col_class_rank}, 年级排名={col_grade_rank}")

        # 从第4行开始读取数据
        for row_idx in range(4, ws.max_row + 1):
            try:
                row = list(ws[row_idx])
                name_cell = row[col_name - 1]
                number_cell = row[col_school_number - 1]
                score_cell = row[col_score - 1]

                # 读取数据
                name = clean_student_name(name_cell.value)
                school_number = str(number_cell.value).strip() if number_cell.value else None

                # 检查是否缺考（成绩为--符号）
                score_value = str(score_cell.value).strip() if score_cell.value else ''
                is_absent = False

                if score_value in ['--', '-', '']:
                    is_absent = True
                    score = None
                else:
                    try:
                        score = float(score_value)
                    except:
                        is_absent = True
                        score = None

                if not name or not school_number:
                    fail_count += 1
                    continue

                # 查找学生ID
                cursor = conn.cursor()
                cursor.execute("SELECT StudentId FROM Students WHERE StudentNumber = ?", (school_number,))
                student_result = cursor.fetchone()

                if not student_result:
                    # 尝试用姓名和班级匹配
                    cursor.execute("SELECT StudentId FROM Students WHERE StudentName = ? AND ClassName = ?",
                                 (name, class_name))
                    student_result = cursor.fetchone()

                if not student_result:
                    errors.append(f"找不到学生: {name}({school_number})")
                    fail_count += 1
                    continue

                student_id = student_result['StudentId']

                # 读取班级排名
                class_rank = None
                class_rank_cell = row[col_class_rank - 1]
                if class_rank_cell.value is not None:
                    rank_value = str(class_rank_cell.value).strip()
                    if rank_value and rank_value not in ['--', '-', '', 'None']:
                        try:
                            # 先尝试int，如果失败再尝试float再转int
                            class_rank = int(float(rank_value))
                        except:
                            try:
                                # 如果float也失败，直接转int
                                class_rank = int(rank_value)
                            except:
                                pass

                # 读取年级排名
                grade_rank = None
                grade_rank_cell = row[col_grade_rank - 1]
                if grade_rank_cell.value is not None:
                    rank_value = str(grade_rank_cell.value).strip()
                    if rank_value and rank_value not in ['--', '-', '', 'None']:
                        try:
                            # 先尝试int，如果失败再尝试float再转int
                            grade_rank = int(float(rank_value))
                        except:
                            try:
                                # 如果float也失败，直接转int
                                grade_rank = int(rank_value)
                            except:
                                pass

                # 检查是否缺考（成绩为--符号）
                score_value = str(score_cell.value).strip() if score_cell.value else ''
                is_absent = False

                if score_value in ['--', '-', '']:
                    is_absent = True
                    score = None
                    # 注意：缺考时排名也要清空
                    class_rank = None
                    grade_rank = None
                else:
                    try:
                        score = float(score_value)
                    except:
                        is_absent = True
                        score = None
                        class_rank = None
                        grade_rank = None

                # 如果缺考，记录但不保存到数据库
                if is_absent:
                    errors.append(f"{name}({school_number}): 缺考")
                    fail_count += 1
                    continue

                # 检查是否已存在（在TimeLimitScores表中使用TimeLimitExamId）
                cursor.execute("""
                    SELECT ScoreId FROM TimeLimitScores
                    WHERE TimeLimitExamId = ? AND StudentId = ? AND SubjectId = ?
                """, (exam_id, student_id, subject_id))

                if cursor.fetchone():
                    # 更新
                    cursor.execute("""
                        UPDATE TimeLimitScores
                        SET Score = ?, ClassRank = ?, GradeRank = ?, UpdatedAt = datetime('now')
                        WHERE TimeLimitExamId = ? AND StudentId = ? AND SubjectId = ?
                    """, (score, class_rank, grade_rank, exam_id, student_id, subject_id))
                else:
                    # 插入
                    cursor.execute("""
                        INSERT INTO TimeLimitScores (TimeLimitExamId, StudentId, SubjectId, Score, ClassRank, GradeRank)
                        VALUES (?, ?, ?, ?, ?, ?)
                    """, (exam_id, student_id, subject_id, score, class_rank, grade_rank))

                conn.commit()
                success_count += 1

            except Exception as e:
                fail_count += 1
                errors.append(f"第{row_idx}行错误: {str(e)}")

    except Exception as e:
        errors.append(f"Sheet读取错误: {str(e)}")

    return success_count, fail_count, errors


def import_time_limit_excel(excel_path, grade_name='高一'):
    """导入限时练Excel文件"""
    print("=" * 80)
    print("限时练成绩导入工具")
    print("=" * 80)

    # 解析考试信息
    filename = excel_path.split('\\')[-1]
    exam_name, subject_name, exam_date = parse_exam_name(filename)

    if not exam_name or not subject_name:
        print(f"❌ 无法从文件名解析考试信息: {filename}")
        return

    print(f"\n考试名称: {exam_name}")
    print(f"科目: {subject_name}")
    print(f"日期: {exam_date}")
    print(f"文件: {filename}")
    print(f"年级: {grade_name}")

    conn = connect_db()

    try:
        # 查找科目ID
        subject_id = find_subject_id(conn, subject_name)
        if not subject_id:
            print(f"❌ 科目不存在: {subject_name}")
            print("   请先在数据库中创建该科目")
            return

        # 创建限时练考试
        exam_id = create_time_limit_exam_if_not_exists(conn, exam_name, subject_name, subject_id, exam_date, grade_name)
        print(f"考试ID: {exam_id}")

        # 加载Excel
        print(f"\n正在加载Excel文件...")
        wb = openpyxl.load_workbook(excel_path, data_only=True)

        total_success = 0
        total_fail = 0
        all_errors = []

        # 遍历所有sheet
        for sheet_name in wb.sheetnames:
            print(f"\n{'='*80}")
            print(f"处理Sheet: {sheet_name}")
            print(f"{'='*80}")

            ws = wb[sheet_name]

            # Sheet名称即为班级名称
            class_name = sheet_name.strip()
            print(f"班级: {class_name}")

            success, fail, errors = import_time_limit_sheet(conn, ws, exam_id, subject_id, class_name)

            total_success += success
            total_fail += fail
            all_errors.extend(errors)

            print(f"\n导入完成: 成功 {success} 条, 失败 {fail} 条")

            # 统计缺考人数
            absent_count = sum(1 for err in errors if '缺考' in err)

            if errors:
                print(f"处理详情:")
                if absent_count > 0:
                    print(f"  缺考: {absent_count} 人")

                # 显示其他错误
                other_errors = [err for err in errors if '缺考' not in err]
                if other_errors:
                    print(f"  错误:")
                    for err in other_errors[:10]:  # 只显示前10个错误
                        print(f"    - {err}")
                    if len(other_errors) > 10:
                        print(f"    ... 还有 {len(other_errors) - 10} 个错误")

        wb.close()

        print(f"\n{'='*80}")
        print(f"总计导入完成: 成功 {total_success} 条, 失败 {total_fail} 条")

        # 统计总缺考人数
        total_absent = sum(1 for err in all_errors if '缺考' in err)
        if total_absent > 0:
            print(f"缺考人数: {total_absent} 人")

        print(f"{'='*80}")

    except Exception as e:
        print(f"\n❌ 发生错误: {e}")
        import traceback
        traceback.print_exc()

    finally:
        conn.close()


def main():
    import sys
    import os

    # 第一次导入
    if len(sys.argv) > 1:
        excel_path = sys.argv[1]
        grade_name = sys.argv[2] if len(sys.argv) > 2 else '高一'
        import_time_limit_excel(excel_path, grade_name)
    else:
        excel_path = input("请输入Excel文件路径: ").strip()
        grade_name = input("请输入年级（默认：高一）: ").strip() or '高一'
        import_time_limit_excel(excel_path, grade_name)

    # 循环导入，直到用户选择退出
    while True:
        print("\n" + "=" * 80)
        print("是否继续导入其他文件？")
        print("  1. 继续导入（拖拽Excel文件到此窗口或输入路径）")
        print("  2. 退出")
        print("=" * 80)

        choice = input("\n请选择（1-2，默认2）: ").strip()
        if not choice or choice == '2':
            print("\n✅ 导入程序结束")
            break

        if choice == '1':
            # 尝试从剪贴板或输入获取文件路径
            print("\n请拖拽Excel文件到此窗口，或输入文件路径:")
            new_excel_path = input().strip()

            # 去除可能包含的引号
            new_excel_path = new_excel_path.strip('"').strip("'")

            if not new_excel_path:
                print("❌ 未提供文件路径，跳过")
                continue

            if not os.path.exists(new_excel_path):
                print(f"❌ 文件不存在: {new_excel_path}")
                continue

            if not new_excel_path.lower().endswith('.xlsx'):
                print("❌ 文件格式错误，请选择.xlsx文件")
                continue

            # 询问是否修改年级
            change_grade = input(f"是否修改年级（当前：{grade_name}）？(y/N): ").strip().lower()
            if change_grade == 'y':
                new_grade = input("请输入年级: ").strip()
                if new_grade:
                    grade_name = new_grade

            # 导入新文件
            import_time_limit_excel(new_excel_path, grade_name)
        else:
            print("❌ 无效选择")



if __name__ == '__main__':
    main()
