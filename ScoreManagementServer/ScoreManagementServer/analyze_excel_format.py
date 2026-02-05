"""分析Excel成绩表格格式"""
from openpyxl import load_workbook

# 读取Excel文件
wb = load_workbook('e:/StudentScoreProject/ScoreManagementServer/ScoreManagementServer/107_学生成绩(方向名次).xlsx')
ws = wb.active

print("="*80)
print("Excel文件分析")
print("="*80)
print(f"\n工作表名: {wb.sheetnames}")
print(f"活动工作表: {ws.title}")
print(f"总行数: {ws.max_row}")
print(f"总列数: {ws.max_column}")

print("\n前10行数据:")
print("-"*80)
for i, row in enumerate(ws.iter_rows(values_only=True, max_row=10)):
    print(f"行{i+1}:")
    for j, cell in enumerate(row):
        if cell is not None:
            print(f"  列{j+1}: {cell}")
    print()

print("\n表头信息（前3行）:")
print("-"*80)
for i in range(1, min(4, ws.max_row + 1)):
    print(f"行{i}:")
    for j in range(1, ws.max_column + 1):
        cell = ws.cell(row=i, column=j).value
        if cell is not None:
            print(f"  列{j}: {cell}")
    print()
