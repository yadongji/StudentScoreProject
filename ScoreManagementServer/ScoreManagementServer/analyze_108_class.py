"""
分析108班成绩统计表.xlsx的结构
"""

from openpyxl import load_workbook
import os
import sys
import io

# 修复Windows控制台编码问题
if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

file_path = "108班成绩统计表.xlsx"

print("=" * 80)
print("分析108班成绩统计表.xlsx")
print("=" * 80)

if not os.path.exists(file_path):
    print(f"文件不存在: {file_path}")
    exit(1)

wb = load_workbook(filename=file_path, read_only=True)
ws = wb.active

print(f"\nExcel文件信息:")
print(f"  工作表数量: {len(wb.sheetnames)}")
print(f"  工作表名称: {wb.sheetnames}")
print(f"  当前工作表: {ws.title}")
print(f"  总行数: {ws.max_row}")
print(f"  总列数: {ws.max_column}")

# 读取前10行的内容
print(f"\n前10行内容:")
print("=" * 80)
for row_idx in range(1, min(11, ws.max_row + 1)):
    row_data = []
    for col_idx in range(1, min(21, ws.max_column + 1)):
        cell_value = ws.cell(row=row_idx, column=col_idx).value
        row_data.append(str(cell_value) if cell_value is not None else "")
    print(f"第{row_idx:2d}行: {' | '.join(row_data[:15])}")  # 只显示前15列

# 读取表头（第1行）
print(f"\n第1行表头:")
headers = [str(ws.cell(row=1, column=col).value) if ws.cell(row=1, column=col).value else "" for col in range(1, ws.max_column + 1)]
for idx, header in enumerate(headers, 1):
    print(f"  列{idx:2d}: {header}")

# 分析每列的数据类型
print(f"\n列数据类型分析:")
print("=" * 80)
for col_idx in range(1, min(21, ws.max_column + 1)):
    header = headers[col_idx - 1]
    print(f"\n列{col_idx} ({header}):")

    # 检查前5行数据
    values = []
    for row_idx in range(2, min(7, ws.max_row + 1)):
        val = ws.cell(row=row_idx, column=col_idx).value
        values.append(val)
        print(f"  第{row_idx}行: {val}")

wb.close()

print("\n" + "=" * 80)
print("分析完成！")
print("=" * 80)
